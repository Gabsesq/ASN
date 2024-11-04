import os
import datetime
from openpyxl.styles import NamedStyle, Alignment
import sys




def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Define shared directories
UPLOAD_FOLDER = resource_path('uploads')
FINISHED_FOLDER = resource_path('Finished')

def format_cells_as_text(worksheet):
    """Format only the cells with data as text."""
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:  # Only apply formatting if there's data
                cell.number_format = "@"

def align_cells_left(worksheet):
    """Align only the cells with data to the left."""
    left_alignment = Alignment(horizontal='left')
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:  # Only align cells with data
                cell.alignment = left_alignment

# Helper function to get the current date as a string
def get_current_date():
    return datetime.datetime.now().strftime("%m.%d.%Y")

# Helper function to create a folder if it doesn't exist
def create_folder(folder_path):
    os.makedirs(folder_path, exist_ok=True)

# Helper function to extract PO number from XLSX or XLS files
def extract_po_number(file_path, is_xlsx=True):
    if is_xlsx:
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        return wb.active['C4'].value
    else:
        import xlrd
        xls_book = xlrd.open_workbook(file_path)
        return xls_book.sheet_by_index(0).cell_value(3, 2)
    
def manyToMany(xls_sheet, source_ws, start_row, start_col, dest_col, dest_start_row, column_length):
    """
    Copies values from the source sheet to the destination sheet, row-by-row.
    Handles cases where only a single row exists.
    """
    print(f"Copying from column {start_col} (starting at row {start_row}) "
          f"to {dest_col}{dest_start_row} for {column_length} rows.")

    for i in range(column_length):
        try:
            value = xls_sheet.cell_value(start_row + i - 1, start_col)
            source_ws[f'{dest_col}{dest_start_row + i}'] = value
            print(f"Pasted '{value}' to {dest_col}{dest_start_row + i}")
        except IndexError as e:
            print(f"IndexError at row {start_row + i - 1}, col {start_col}: {str(e)}")
        except Exception as e:
            print(f"Unexpected error: {str(e)}")
# ExcelHelpers.py

def oneToMany(xls_sheet, source_ws, row, col, target_column, start_row, column_length):
    """
    Copies a value from one specific cell and pastes it into multiple rows in a target column.
    """
    try:
        # Fetch the value from the specific cell
        value = xls_sheet.cell_value(row, col)
        print(f"Copying value '{value}' from ({row + 1}, {col + 1})")

        # Ensure exactly `column_length` rows are pasted without overshooting
        for i in range(column_length):
            current_row = start_row + i  # Adjust to paste into the correct row
            source_ws[f'{target_column}{current_row}'] = value
            print(f"Pasting '{value}' into {target_column}{current_row}")

    except IndexError as e:
        print(f"Error accessing cell ({row}, {col}): {str(e)}")
    except Exception as e:
        print(f"Unexpected error: {str(e)}")




def typedValue(source_ws, static_value, target_column, start_row, column_length):
    """
    Pastes a static value (like "N/A") into multiple rows in a target column.

    Parameters:
        source_ws: The destination worksheet object.
        static_value: The value to paste (e.g., "N/A").
        target_column (str): The column letter in the destination sheet (e.g., 'D').
        start_row (int): The starting row in the destination sheet (e.g., 19).
        column_length (int): Number of rows to paste the value into.
    """
    try:
        print(f"Using static value '{static_value}'")

        # Paste the static value into the specified column for the given length
        for i in range(start_row, start_row + column_length):
            source_ws[f'{target_column}{i}'] = static_value
            print(f"Pasting '{static_value}' into {target_column}{i}")

    except Exception as e:
        print(f"Unexpected error: {str(e)}")

def get_column_length(sheet, start_row):
    """Calculate the number of non-empty rows, ensuring at least one row is processed."""
    column_length = 0
    while True:
        try:
            value = sheet.cell_value(start_row - 1, 0)  # Column A (index 0)
            print(f"Row {start_row}: Value in A = '{value}'")
            if value:
                column_length += 1
                start_row += 1
            else:
                break
        except IndexError as e:
            print(f"IndexError accessing row {start_row - 1}, column 0: {str(e)}")
            break

    # Ensure at least one row is processed, even if only one row exists
    if column_length == 0:
        column_length = 1
        print("Column length adjusted to 1 to handle single-row data.")

    print(f"Final Column Length: {column_length}")
    return column_length



def QTY_total(sheet, start_row, qty_column):
    """
    Calculate the total quantity starting from a specific row and column.

    Parameters:
        sheet: The Excel sheet object (can be openpyxl or xlrd sheet).
        start_row (int): The row to start processing (1-based index).
        qty_column (int): The column containing quantity values (0-based index).

    Returns:
        int: Total quantity sum.
    """
    total_quantity = 0  # Initialize the total quantity

    print(f"Calculating quantity total starting from row {start_row}, column {qty_column}...")

    row = start_row
    while True:
        try:
            # Read the value from the sheet
            value = sheet.cell_value(row - 1, qty_column)  # Adjust to 0-based index
            print(f"Row {row}: Raw Quantity Value = '{value}'")

            # Convert to integer (if numeric) and add to the total
            if isinstance(value, str) and value.strip().isdigit():
                value = int(value)
            elif isinstance(value, float):
                value = int(value)

            if isinstance(value, int):
                total_quantity += value
                print(f"Added {value} to total. Current Total: {total_quantity}")
            else:
                print(f"Skipping non-numeric value at row {row}: '{value}'")

            row += 1  # Move to the next row

        except IndexError:
            print(f"Reached the end of the data at row {row}.")
            break
        except ValueError as e:
            print(f"ValueError at row {row}: {str(e)}")
            break

    print(f"Final Quantity Total: {total_quantity}")
    return total_quantity


def generate_rows(sheet, start_row, qty_column, column_count):
    """
    Generate rows based on the quantity value in a specified column.
    Each duplicated row will maintain the same content but with unique line numbers.

    Parameters:
        sheet: The Excel sheet object (can be openpyxl or xlrd sheet).
        start_row (int): The row to start reading from (1-based index).
        qty_column (int): The column containing quantity values (0-based index).
        column_count (int): The number of columns to copy for each row.

    Returns:
        list: A list of rows where each row is a list of cell values, 
              with the first column representing the correct line number.
    """
    generated_rows = []  # Store generated rows with correct line numbering
    line_number = 1  # Start the line number from 1

    row = start_row
    while True:
        try:
            # Read the quantity value from the specified column
            qty_value = sheet.cell_value(row - 1, qty_column)
            qty = int(float(qty_value)) if qty_value else 1  # Handle empty or non-numeric values

            # Extract the row data starting from column 0 to `column_count`
            row_data = [sheet.cell_value(row - 1, col) for col in range(column_count)]

            # Duplicate the row based on the quantity value
            for _ in range(qty):
                # Add the current line number as the first element in the row
                generated_rows.append([line_number] + row_data)
                print(f"Generated row: {generated_rows[-1]}")
                line_number += 1  # Increment line number for each new row

            row += 1  # Move to the next row

        except IndexError:
            print(f"Reached the end of the sheet at row {row}.")
            break
        except ValueError as e:
            print(f"ValueError at row {row}: {e}")
            break

    return generated_rows

