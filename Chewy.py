from flask import Flask, request, jsonify, render_template
from openpyxl import load_workbook, Workbook
import os
import xlrd
import datetime

app = Flask(__name__)

# Create upload folder
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Define source files and destination copies
source_asn_xlsx = "assets/Chewy/Chewy 856 ASN - Copy.xlsx"
source_label_xls = "assets/Chewy/Chewy UCC128 Label Request - Copy.xls"
asn_copy_backup = "Finished/Chewy/Chewy 856 ASN - Copy - Backup.xlsx"
label_copy_backup = "Finished/Chewy/Chewy UCC128 Label Request - Copy - Backup.xlsx"

# Function to copy data from uploaded .xlsx file to specific cells in the ASN .xlsx backup
def copy_xlsx_data(uploaded_file, dest_file):
    print("copy_xlsx_data function called!")  # Debug print to check if function is called

    uploaded_wb = load_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    uploaded_ws = uploaded_wb.active
    source_ws = source_wb.active

    # Mapping uploaded cells to copy cells
    data_map = {
        'B16': 'E3',   # uploaded B16 -> copy E3
        'D16': 'E4',   # uploaded D16 -> copy E4
        'E16': 'E5',   # uploaded E16 -> copy E5
        'J16': 'E6',   # uploaded J16 -> copy E6
        'K16': 'E7',   # uploaded K16 -> copy E7
        'L16': 'E8',   # uploaded L16 -> copy E8
        'C8':  'E14',  # uploaded C8  -> copy E14
    }

    # Loop through the map and transfer data from the uploaded file to the backup copy
    for upload_cell, copy_cell in data_map.items():
        source_ws[copy_cell] = uploaded_ws[upload_cell].value

    # Track numbers from A21 and below, copy to A20 in the copy
    row = 21
    data_to_copy = []
    while uploaded_ws[f'A{row}'].value is not None:
        data_to_copy.append(uploaded_ws[f'A{row}'].value)
        row += 1

    # Paste data starting from A20 in the backup copy
    for i, value in enumerate(data_to_copy):
        source_ws[f'A{20 + i}'] = value

    # NEW: Dynamically determine range based on column A data
    count = len(data_to_copy)  # Get the count of rows from column A

    # Get value from C4 in the uploaded file and print it for debugging
    value_from_upload = uploaded_ws['C4'].value
    print(f"Value from C4 in uploaded file: {value_from_upload}")  # Print the value for verification

    # Copy the value from C4 to B20 through B(dynamic count based on A column)
    if value_from_upload is not None:
        for i in range(20, 20 + count):  # Dynamic range based on column A length
            source_ws[f'B{i}'] = value_from_upload
    else:
        print("No value found in C4 or unable to read the value.")  # Debugging if C4 is None

    # Save the updated copy
    source_wb.save(dest_file)


# Function to convert .xls to .xlsx and transfer data to a backup of ASN copy
def convert_xls_data(uploaded_file, dest_file):
    print("convert_xls_data function called!")  # Debugging print

    # Open .xls file using xlrd
    xls_book = xlrd.open_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    source_ws = source_wb.active
    xls_sheet = xls_book.sheet_by_index(0)

    # Mapping uploaded cells to copy cells
    data_map = {
        (15, 1): 'E3',   # uploaded B16 -> copy E3 (zero-indexed for xlrd)
        (15, 3): 'E4',   # uploaded D16 -> copy E4
        (15, 4): 'E5',   # uploaded E16 -> copy E5
        (15, 9): 'E6',   # uploaded J16 -> copy E6
        (15, 10): 'E7',  # uploaded K16 -> copy E7
        (15, 11): 'E8',  # uploaded L16 -> copy E8
        (7, 2): 'E14',   # uploaded C8 -> copy E14
    }

    # Transfer data from .xls to the backup copy
    for (row, col), copy_cell in data_map.items():
        value = xls_sheet.cell_value(row, col)
        print(f"Value from row {row+1}, col {col+1} (uploaded file): {value}")  # Debugging print
        source_ws[copy_cell] = value

    # Get the value from C4 (row 3, column 2 in zero-based indexing)
    value_from_upload_C4 = xls_sheet.cell_value(3, 2)  # C4 in zero-based index (row 3, column 2)
    print(f"Value from C4 in .xls file: {value_from_upload_C4}")  # Debugging print

    # Get the value from H4 (row 3, column 7 in zero-based indexing) - this is the date
    value_from_upload_H4 = xls_sheet.cell_value(3, 7)  # H4 in zero-based index (row 3, column 7)
    print(f"Date value from H4 in .xls file: {value_from_upload_H4}")  # Debugging print

    # Track numbers from A21 and below, copy to A20 in the copy
    row = 21  # Start from row 21 in the uploaded file
    copy_row = 20  # Start pasting from row 20 in the destination file
    column_length = 0  # Variable to store the length of column A

    while True:
        try:
            value = xls_sheet.cell_value(row - 1, 0)  # Column A is index 0 (zero-based)
            if value:
                source_ws[f'A{copy_row}'] = value
                row += 1
                copy_row += 1
                column_length += 1  # Increment the length count
            else:
                break
        except IndexError:
            break

    print(f"Length of Column A: {column_length}")  # Debugging print for column length

    # Now copy from upload file column H (starting at H21) down to E20 in the copy file
    for i in range(21, 21 + column_length):  # Loop through rows in column H in the upload file
        value_from_H = xls_sheet.cell_value(i - 1, 7)  # H21 is (row 20, col 7) in zero-based indexing
        source_ws[f'E{i - 1}'] = value_from_H  # Paste into E20 and down
        print(f"Pasting {value_from_H} from H{i} to E{i - 1}")  # Debugging print

    # Now copy value from C4 into B column (B20 and down for column_length rows)
    if value_from_upload_C4 is not None:
        for i in range(20, 20 + column_length):  # Copy value to B20 through B(dynamic based on column A length)
            source_ws[f'B{i}'] = value_from_upload_C4
            print(f"Pasting {value_from_upload_C4} into B{i}")  # Debugging print
    else:
        print("No value found in C4 or unable to read the value.")  # Debugging if C4 is None

    # Copy date from H4 into C column (C20 and down for column_length rows)
    if value_from_upload_H4 is not None:
        for i in range(20, 20 + column_length):  # Copy value to C20 through C(dynamic based on column A length)
            source_ws[f'C{i}'] = value_from_upload_H4
            print(f"Pasting date {value_from_upload_H4} into C{i}")  # Debugging print
    else:
        print("No date found in H4 or unable to read the value.")  # Debugging if H4 is None

    # Now copy from upload file column G (starting at G21) down to F20 in the copy file
    for i in range(21, 21 + column_length):  # Loop through rows in column G in the upload file
        value_from_G = xls_sheet.cell_value(i - 1, 6)  # G21 is (row 20, col 6) in zero-based indexing
        source_ws[f'F{i - 1}'] = value_from_G  # Paste into F20 and down
        print(f"Pasting {value_from_G} from G{i} to F{i - 1}")  # Debugging print

    # Now copy from upload file column I (starting at I21) down to G20 in the copy file
    for i in range(21, 21 + column_length):  # Loop through rows in column I in the upload file
        value_from_I = xls_sheet.cell_value(i - 1, 8)  # I21 is (row 20, col 8) in zero-based indexing
        source_ws[f'G{i - 1}'] = value_from_I  # Paste into G20 and down
        print(f"Pasting {value_from_I} from I{i} to G{i - 1}")  # Debugging print

    # Now copy from upload file column B (starting at B21) down to H20 in the copy file
    for i in range(21, 21 + column_length):  # Loop through rows in column B in the upload file
        value_from_B = xls_sheet.cell_value(i - 1, 1)  # B21 is (row 20, col 1) in zero-based indexing
        source_ws[f'H{i - 1}'] = value_from_B  # Paste into H20 and down
        print(f"Pasting {value_from_B} from B{i} to H{i - 1}")  # Debugging print

    # Now copy from upload file column C (starting at C21) down to I20 in the copy file
    for i in range(21, 21 + column_length):  # Loop through rows in column C in the upload file
        value_from_C = xls_sheet.cell_value(i - 1, 2)  # C21 is (row 20, col 2) in zero-based indexing
        source_ws[f'I{i - 1}'] = value_from_C  # Paste into I20 and down
        print(f"Pasting {value_from_C} from C{i} to I{i - 1}")  # Debugging print

    # Now copy from upload file column E (starting at E21) down to L20 in the copy file
    for i in range(21, 21 + column_length):  # Loop through rows in column E in the upload file
        value_from_E = xls_sheet.cell_value(i - 1, 4)  # E21 is (row 20, col 4) in zero-based indexing
        source_ws[f'L{i - 1}'] = value_from_E  # Paste into L20 and down
        print(f"Pasting {value_from_E} from E{i} to L{i - 1}")  # Debugging print

    
    # Save the updated copy
    source_wb.save(dest_file)



# Serve the upload form on the root path
@app.route('/')
def upload_form():
    return '''
    <h2>Upload an Excel File</h2>
    <form id="uploadForm" enctype="multipart/form-data" method="POST" action="/upload">
        <input type="file" id="excelFile" name="file" accept=".xlsx, .xls" required>
        <button type="submit">Upload and Transfer Data</button>
    </form>
    '''

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'message': 'No file part'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'message': 'No selected file'}), 400

    if file:
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(file_path)

        try:
            if file.filename.endswith('.xlsx'):
                print("Processing .xlsx file")  # New debug print
                # Get the value from C4 and the current date
                uploaded_wb = load_workbook(file_path)
                uploaded_ws = uploaded_wb.active
                po_number = uploaded_ws['C4'].value  # Get the value from C4
                current_date = datetime.datetime.now().strftime("%m.%d.%Y")  # Get the current date
                # Create dynamic filename for the backup
                asn_copy_backup = f"Finished/Chewy/Chewy 856 ASN PO {po_number} {current_date}.xlsx"
                print(f"Generated file name: {asn_copy_backup}")  # Debugging

                # Copy the data to the backup of the ASN .xlsx file
                copy_xlsx_data(file_path, asn_copy_backup)

            elif file.filename.endswith('.xls'):
                print("Processing .xls file")  # New debug print
                # Get the value from C4 and the current date
                xls_book = xlrd.open_workbook(file_path)
                xls_sheet = xls_book.sheet_by_index(0)
                po_number = xls_sheet.cell_value(3, 2)  # Get value from C4 in the uploaded .xls file
                current_date = datetime.datetime.now().strftime("%m.%d.%Y")  # Get the current date
                # Create dynamic filename for the backup
                asn_copy_backup = f"Finished/Chewy/Chewy 856 ASN PO {po_number} {current_date}.xlsx"
                print(f"Generated file name: {asn_copy_backup}")  # Debugging

                # Convert the .xls and transfer data to the backup of the label .xls file
                convert_xls_data(file_path, asn_copy_backup)

        except Exception as e:
            print(f"Error: {e}")  # New debug print for errors
            return jsonify({'message': f'Error processing file: {str(e)}'}), 500

        return jsonify({'message': 'File processed and data transferred successfully!'})

if __name__ == '__main__':
    app.run(debug=True)
