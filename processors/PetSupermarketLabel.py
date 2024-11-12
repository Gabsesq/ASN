import os
import datetime
from openpyxl import load_workbook
import xlrd
from ExcelHelpers import (
    resource_path, FINISHED_FOLDER, format_cells_as_text, align_cells_left
)


# Define source files and destination copies for Pet Supermarket
source_label_xlsx = resource_path("assets/Pet Supermarket/Blank Pet Supermarket UCC 128 Label Request.xlsx")

# Function to copy data from uploaded .xlsx file to specific cells in the label request .xlsx backup
def copy_xlsx_data(uploaded_file, dest_file):
    uploaded_wb = load_workbook(uploaded_file)
    source_wb = load_workbook(source_label_xlsx)
    uploaded_ws = uploaded_wb.active
    source_ws = source_wb.active

    # Mapping uploaded cells to copy cells
    data_map = {
        'C4': 'A14',  # Copy C4 to A14
        'L12': 'B14',  # Copy L12 to B14
    }

    # Transfer data from the uploaded file to the backup copy
    for upload_cell, copy_cell in data_map.items():
        source_ws[copy_cell] = uploaded_ws[upload_cell].value

    # Set static values
    source_ws['C14'] = "SAIA"
    source_ws['E14'] = "mixed"
    source_ws['G14'] = "mixed"
    source_ws['H14'] = 1

    # Save the updated copy
    source_wb.save(dest_file)

# Function to convert .xls to .xlsx and transfer data to a backup of the label request copy
def convert_xls_data(uploaded_file, dest_file):
    xls_book = xlrd.open_workbook(uploaded_file)
    source_wb = load_workbook(source_label_xlsx)
    source_ws = source_wb.active
    xls_sheet = xls_book.sheet_by_index(0)

    # Mapping uploaded cells to copy cells
    data_map = {
        (3, 2): 'A14',  # C4 is (row 3, col 2) in zero-based indexing, copying to A14
        (11, 11): 'B14'  # L12 is (row 11, col 11) in zero-based indexing, copying to B14
    }

    # Transfer data from the uploaded .xls file to the backup copy
    for (row, col), copy_cell in data_map.items():
        value = xls_sheet.cell_value(row, col)
        source_ws[copy_cell] = value

    # Set static values
    source_ws['C14'] = "SAIA"
    source_ws['E14'] = "mixed"
    source_ws['G14'] = "mixed"
    source_ws['H14'] = 1

    format_cells_as_text(source_ws)
    align_cells_left(source_ws)
    align_cells_left(source_ws)
    # Save the updated copy
    source_wb.save(dest_file)

def process_PetSupermarketLabel(file_path):
    """Main function to process Pet Supermarket Label Request files."""
    current_date = datetime.datetime.now().strftime("%m.%d.%Y")

    # Determine if the file is XLSX or XLS and extract the PO number
    if file_path.endswith('.xlsx'):
        uploaded_wb = load_workbook(file_path)
        po_number = uploaded_wb.active['C4'].value
    elif file_path.endswith('.xls'):
        xls_book = xlrd.open_workbook(file_path)
        po_number = xls_book.sheet_by_index(0).cell_value(3, 2)

    # Define the backup file path in FINISHED_FOLDER with a PetSupermarket subfolder
    backup_file = os.path.join(FINISHED_FOLDER, f"PetSupermarket/Pet Supermarket Label Request PO {po_number} {current_date}.xlsx")
    
    # Ensure the directory exists
    os.makedirs(os.path.dirname(backup_file), exist_ok=True)

    # Perform the copy or conversion based on file type
    if file_path.endswith('.xlsx'):
        copy_xlsx_data(file_path, backup_file)
    elif file_path.endswith('.xls'):
        convert_xls_data(file_path, backup_file)

    return backup_file, po_number
