from openpyxl import load_workbook
import xlrd
import datetime
import os
import sys
import tempfile
from ExcelHelpers import (
    QTY_total, resource_path, oneToMany, manyToMany, get_current_date, extract_po_number, format_cells_as_text, align_cells_left, get_column_length, FINISHED_FOLDER
)
from upc_counts import counts  # Add this import



# Define source file for TSC IS
source_asn_xls = resource_path("assets/TSCIS/Master Template Tractor Supply IS ASN.xlsx")

# Add debug print
print(f"Looking for template at: {source_asn_xls}")
if not os.path.exists(source_asn_xls):
    print(f"WARNING: Template file not found at {source_asn_xls}")
else:
    print(f"Template file found at {source_asn_xls}")


# Function to copy data from uploaded .xlsx file to the ASN .xlsx backup
def copy_xlsx_data(uploaded_file, dest_file):
    try:
        # For xlsx files, we'll convert the template first
        output_wb = load_workbook(source_asn_xls.replace('.xls', '.xlsx'))
        uploaded_wb = load_workbook(uploaded_file)
        
        output_ws = output_wb.active
        uploaded_ws = uploaded_wb.active

        format_cells_as_text(output_ws)
        align_cells_left(output_ws)

        # Mapping uploaded cells to copy cells
        data_map = {
            'B14': 'E3', 'D14': 'E4', 'E14': 'E5',
            'J14': 'E6', 'K14': 'E7', 'L14': 'E8',
            'C9': 'B11'
        }

        # Transfer data with debugging
        for upload_cell, copy_cell in data_map.items():
            try:
                value = uploaded_ws[upload_cell].value
                output_ws[copy_cell] = value
                print(f"Copied value '{value}' from {upload_cell} to {copy_cell}")
            except Exception as e:
                print(f"Error copying from {upload_cell} to {copy_cell}: {str(e)}")

        # Calculate dynamic column length starting from A18
        column_length = get_column_length(uploaded_ws, start_row=18)

        # Copy value from 'C4' into 'B17' and down using oneToMany
        oneToMany(uploaded_ws, output_ws, 3, 2, 'B', 17, column_length)

        format_cells_as_text(output_ws)
        align_cells_left(output_ws)
        align_cells_left(output_ws)
        output_wb.save(dest_file)
        print(f"Saved file successfully as {dest_file}")
        return True

    except Exception as e:
        print(f"Error in copy_xlsx_data: {str(e)}")
        return False

# Function to convert and copy data from .xls files
def convert_xls_data(uploaded_file, dest_file):
    try:
        source_wb = load_workbook(source_asn_xls)
        source_ws = source_wb.active
        
        xls_book = xlrd.open_workbook(uploaded_file)
        xls_sheet = xls_book.sheet_by_index(0)

        format_cells_as_text(source_ws)
        align_cells_left(source_ws)

        # Debug print to see sheet dimensions
        print(f"Sheet dimensions: rows={xls_sheet.nrows}, cols={xls_sheet.ncols}")

        # Mapping uploaded cells to copy cells with error handling
        data_map = {
            (13, 1): 'E4', (13, 3): 'E5', (13, 4): 'E6',
            (13, 9): 'E7', (13, 10): 'E8', (13, 11): 'E9',
            (6,2) : 'E11'
        }
        
        for (row, col), copy_cell in data_map.items():
            try:
                value = xls_sheet.cell_value(row, col)
                source_ws[copy_cell] = value
                print(f"Copied value '{value}' from ({row}, {col}) to {copy_cell}")
            except IndexError as e:
                print(f"IndexError: ({row}, {col}) is out of range: {str(e)}")
            except Exception as e:
                print(f"Unexpected error copying ({row}, {col}) to {copy_cell}: {str(e)}")

        # --- AGGREGATE BY PRODUCT (UPC) ---
        product_dict = {}  # key: UPC, value: dict with total_qty and other info
        item_start_row = 17
        while True:
            try:
                qty = int(float(xls_sheet.cell_value(item_start_row, 1)))  # Column B (QTY)
                upc = str(int(xls_sheet.cell_value(item_start_row, 4)))  # Column E (UPC)
                description = xls_sheet.cell_value(item_start_row, 7)  # Column H (Description)
                vendor_part = xls_sheet.cell_value(item_start_row, 6)  # Column G (Vendor Part)
                buyer_part = xls_sheet.cell_value(item_start_row, 5)  # Column F (Buyer Part)
                uom = xls_sheet.cell_value(item_start_row, 2)  # Column C (UOM)
                po_number = xls_sheet.cell_value(3, 2)  # PO Number

                if upc in product_dict:
                    product_dict[upc]['qty'] += qty
                else:
                    product_dict[upc] = {
                        'qty': qty,
                        'description': description,
                        'vendor_part': vendor_part,
                        'buyer_part': buyer_part,
                        'uom': uom,
                        'po_number': po_number
                    }

                item_start_row += 1
                if item_start_row >= xls_sheet.nrows or not xls_sheet.cell_value(item_start_row, 1):
                    break
            except IndexError as e:
                print(f"IndexError at row {item_start_row}: {str(e)}")
                break
            except Exception as e:
                print(f"Error processing row {item_start_row}: {str(e)}")
                break

        # --- WRITE ONE LINE PER PRODUCT ---
        output_row = 17
        line_number = 1
        for upc, info in product_dict.items():
            source_ws[f'A{output_row}'] = line_number  # Line number
            source_ws[f'B{output_row}'] = info['po_number']  # PO Number
            source_ws[f'D{output_row}'] = "NA"  # Description
            source_ws[f'E{output_row}'] = upc  # UPC
            source_ws[f'F{output_row}'] = info['buyer_part']  # Buyer Part
            source_ws[f'G{output_row}'] = info['vendor_part']  # Vendor Part
            source_ws[f'H{output_row}'] = info['qty']  # TOTAL QTY for this product
            source_ws[f'I{output_row}'] = info['uom']  # UOM
            source_ws[f'L{output_row}'] = info['description']  # Description
            output_row += 1
            line_number += 1

        # Update total quantity (sum of all products)
        source_ws['B13'] = 1
        print(f"Set B13 to 1 as requested.")

        format_cells_as_text(source_ws)
        align_cells_left(source_ws)

        source_wb.save(dest_file)
        print(f"Saved file successfully as {dest_file}")
        return True

    except Exception as e:
        print(f"Error in convert_xls_data: {str(e)}")
        return False


def process_TSCISASN(file_path):
    """Main function to process TSCIS ASN files."""
    try:
        current_date = get_current_date()
        po_number = extract_po_number(file_path, is_xlsx=file_path.endswith('.xlsx'))
        
        # Fix the path separator issue by using os.path.join
        backup_file = os.path.join(FINISHED_FOLDER, 'TSCIS', f"Tractor Supply IS ASN {po_number} {current_date}.xlsx")
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(backup_file), exist_ok=True)

        success = False
        if file_path.endswith('.xlsx'):
            success = copy_xlsx_data(file_path, backup_file)
        elif file_path.endswith('.xls'):
            success = convert_xls_data(file_path, backup_file)
        
        if not success:
            raise Exception("Failed to process file")
            
        if not os.path.exists(backup_file):
            raise Exception(f"Output file was not created at {backup_file}")
            
        return backup_file, po_number

    except Exception as e:
        print(f"Error in process_TSCISASN: {str(e)}")
        raise


