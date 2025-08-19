from openpyxl import load_workbook
import xlrd
import datetime
import os
from ExcelHelpers import (
    resource_path, FINISHED_FOLDER, format_cells_as_text, align_cells_left, manyToMany, oneToMany, typedValue
)
from upc_counts import counts  # Import the UPC counts dictionary

# Define source files and destination copies for Chewy
source_asn_xlsx = resource_path("assets/Murdochs/Blank Murdochs 856 ASN.xlsx")

# Function to copy data from uploaded .xlsx file to specific cells in the ASN .xlsx backup
def copy_xlsx_data(uploaded_file, dest_file):
    uploaded_wb = load_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    uploaded_ws = uploaded_wb.active
    source_ws = source_wb.active
    format_cells_as_text(source_ws)
    align_cells_left(source_ws)
    # Mapping uploaded cells to copy cells
    data_map = {
        'B18': 'E3', 'D18': 'E4', 'E18': 'E5',
        'J18': 'E6', 'K18': 'E7', 'L18': 'E8',
        'C10': 'E14'
    }

    # Transfer data from the uploaded file to the backup copy
    for upload_cell, copy_cell in data_map.items():
        source_ws[copy_cell] = uploaded_ws[upload_cell].value

    # Manually assign values for cells that don't come from the uploaded file
    source_ws['B12'] = "FEDG"
    source_ws['B13'] = "Fedex"
    
    # Set current date in E11 and highlight it
    current_date = datetime.datetime.now().strftime("%m/%d/%Y")
    source_ws['E11'] = current_date
    
    # Highlight cell E11 with yellow background
    from openpyxl.styles import PatternFill
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    source_ws['E11'].fill = yellow_fill

    # Create carton-based lines from the uploaded Excel file
    carton_lines_count = create_carton_based_lines_from_xlsx(uploaded_ws, source_ws, start_row=23, target_start_row=19)
    
    print(f"Total carton lines created: {carton_lines_count}")
    
    # Additional copy and paste operations with helper functions
    oneToMany_xlsx(uploaded_ws, source_ws, row=4, col=2, target_column='B', start_row=19, column_length=carton_lines_count)  # PO
    oneToMany_xlsx(uploaded_ws, source_ws, row=4, col=7, target_column='C', start_row=19, column_length=carton_lines_count)  # PO Date

    typedValue(source_ws, static_value="NA", target_column='D', start_row=19, column_length=carton_lines_count)

    # Set total cartons in B15 and "C" in E13
    source_ws['B15'] = carton_lines_count
    source_ws['E13'] = "C"

    source_wb.save(dest_file)

def create_carton_based_lines_from_xlsx(uploaded_ws, source_ws, start_row=23, target_start_row=19):
    """
    Creates one line per carton instead of one line per product for XLSX files.
    For each product, calculates how many cartons are needed and creates that many lines.
    """
    carton_lines = []  # Store all the data for carton-based lines
    
    # Process each product row from the uploaded file
    row = start_row
    while uploaded_ws[f'A{row}'].value is not None:
        try:
            # Get product data from the uploaded file
            item_no = uploaded_ws[f'A{row}'].value  # Column A - Item number
            qty = int(uploaded_ws[f'B{row}'].value)  # Column B - QTY
            uom = uploaded_ws[f'C{row}'].value  # Column C - Unit of Measure
            unit_price = uploaded_ws[f'D{row}'].value  # Column D - Unit Price
            description = uploaded_ws[f'E{row}'].value  # Column E - Description
            upc = uploaded_ws[f'F{row}'].value  # Column F - UPC
            vendor_part = uploaded_ws[f'G{row}'].value  # Column G - Vendor Part
            sku = uploaded_ws[f'I{row}'].value  # Column I - SKU
            
            # Skip if no item number (empty row)
            if not item_no:
                row += 1
                continue
                
            # Calculate number of cartons needed
            if upc in counts:
                items_per_case = counts[upc]
                cartons_needed = qty // items_per_case
                remainder = qty % items_per_case
                
                print(f"Product {item_no}: QTY={qty}, Items per case={items_per_case}, Cartons={cartons_needed}, Remainder={remainder}")
                
                # Create one line per full carton
                for carton_num in range(cartons_needed):
                    carton_lines.append({
                        'item_no': item_no,
                        'qty': items_per_case,  # Full carton quantity
                        'uom': uom,
                        'unit_price': unit_price,
                        'description': description,
                        'vendor_part': vendor_part,
                        'sku': sku,
                        'upc': upc
                    })
                
                # If there's a remainder, create one more line for the partial carton
                if remainder > 0:
                    carton_lines.append({
                        'item_no': item_no,
                        'qty': remainder,  # Partial carton quantity
                        'uom': uom,
                        'unit_price': unit_price,
                        'description': description,
                        'vendor_part': vendor_part,
                        'sku': sku,
                        'upc': upc
                    })
            else:
                print(f"Warning: UPC {upc} not found in counts dictionary for item {item_no}")
                # If UPC not found, create one line with original quantity
                carton_lines.append({
                    'item_no': item_no,
                    'qty': qty,
                    'uom': uom,
                    'unit_price': unit_price,
                    'description': description,
                    'vendor_part': vendor_part,
                    'sku': sku,
                    'upc': upc
                })
            
            row += 1
                
        except (ValueError, TypeError) as e:
            print(f"Error processing row {row}: {e}")
            row += 1
            continue
    
    # Write the carton-based lines to the destination worksheet
    for i, line_data in enumerate(carton_lines):
        target_row = target_start_row + i
        
        # Write data to the appropriate columns
        source_ws[f'A{target_row}'] = line_data['item_no']  # Item number
        source_ws[f'F{target_row}'] = line_data['upc']      # UPC
        source_ws[f'G{target_row}'] = line_data['sku']      # SKU
        source_ws[f'H{target_row}'] = line_data['vendor_part']  # Vendor Part
        source_ws[f'I{target_row}'] = line_data['qty']      # QTY (now per carton)
        source_ws[f'J{target_row}'] = line_data['uom']      # Unit of Measure
        source_ws[f'K{target_row}'] = line_data['description']  # Description
        
        print(f"Created carton line {i+1}: Item {line_data['item_no']}, QTY {line_data['qty']}, UPC {line_data['upc']}")
    
    return len(carton_lines)  # Return the number of carton lines created

def create_carton_based_lines(xls_sheet, source_ws, start_row=23, target_start_row=19):
    """
    Creates one line per carton instead of one line per product for XLS files.
    For each product, calculates how many cartons are needed and creates that many lines.
    """
    carton_lines = []  # Store all the data for carton-based lines
    
    # Process each product row from the uploaded file
    for row_idx in range(start_row - 1, xls_sheet.nrows):  # start_row is 1-indexed, so subtract 1
        try:
            # Get product data from the uploaded file
            item_no = xls_sheet.cell_value(row_idx, 0)  # Column A - Item number
            qty = int(xls_sheet.cell_value(row_idx, 1))  # Column B - QTY
            uom = xls_sheet.cell_value(row_idx, 2)  # Column C - Unit of Measure
            unit_price = xls_sheet.cell_value(row_idx, 3)  # Column D - Unit Price
            description = xls_sheet.cell_value(row_idx, 4)  # Column E - Description
            vendor_part = xls_sheet.cell_value(row_idx, 6)  # Column G - Vendor Part
            sku = xls_sheet.cell_value(row_idx, 8)  # Column I - SKU
            upc = xls_sheet.cell_value(row_idx, 5)  # Column F - UPC
            
            # Skip if no item number (empty row)
            if not item_no:
                continue
                
            # Calculate number of cartons needed
            if upc in counts:
                items_per_case = counts[upc]
                cartons_needed = qty // items_per_case
                remainder = qty % items_per_case
                
                print(f"Product {item_no}: QTY={qty}, Items per case={items_per_case}, Cartons={cartons_needed}, Remainder={remainder}")
                
                # Create one line per full carton
                for carton_num in range(cartons_needed):
                    carton_lines.append({
                        'item_no': item_no,
                        'qty': items_per_case,  # Full carton quantity
                        'uom': uom,
                        'unit_price': unit_price,
                        'description': description,
                        'vendor_part': vendor_part,
                        'sku': sku,
                        'upc': upc
                    })
                
                # If there's a remainder, create one more line for the partial carton
                if remainder > 0:
                    carton_lines.append({
                        'item_no': item_no,
                        'qty': remainder,  # Partial carton quantity
                        'uom': uom,
                        'unit_price': unit_price,
                        'description': description,
                        'vendor_part': vendor_part,
                        'sku': sku,
                        'upc': upc
                    })
            else:
                print(f"Warning: UPC {upc} not found in counts dictionary for item {item_no}")
                # If UPC not found, create one line with original quantity
                carton_lines.append({
                    'item_no': item_no,
                    'qty': qty,
                    'uom': uom,
                    'unit_price': unit_price,
                    'description': description,
                    'vendor_part': vendor_part,
                    'sku': sku,
                    'upc': upc
                })
                
        except (ValueError, IndexError) as e:
            print(f"Error processing row {row_idx + 1}: {e}")
            continue
    
    # Write the carton-based lines to the destination worksheet
    for i, line_data in enumerate(carton_lines):
        target_row = target_start_row + i
        
        # Write data to the appropriate columns
        source_ws[f'A{target_row}'] = line_data['item_no']  # Item number
        source_ws[f'F{target_row}'] = line_data['upc']      # UPC
        source_ws[f'G{target_row}'] = line_data['sku']      # SKU
        source_ws[f'H{target_row}'] = line_data['vendor_part']  # Vendor Part
        source_ws[f'I{target_row}'] = line_data['qty']      # QTY (now per carton)
        source_ws[f'J{target_row}'] = line_data['uom']      # Unit of Measure
        source_ws[f'K{target_row}'] = line_data['description']  # Description
        
        print(f"Created carton line {i+1}: Item {line_data['item_no']}, QTY {line_data['qty']}, UPC {line_data['upc']}")
    
    return len(carton_lines)  # Return the number of carton lines created

def oneToMany_xlsx(uploaded_ws, source_ws, row, col, target_column, start_row, column_length):
    """
    Copies a value from one specific cell in XLSX and pastes it into multiple rows in a target column.
    """
    try:
        # Convert 1-indexed to 0-indexed for openpyxl
        cell_value = uploaded_ws.cell(row=row, column=col+1).value
        print(f"Copying value '{cell_value}' from ({row}, {col + 1})")

        # Ensure exactly `column_length` rows are pasted without overshooting
        for i in range(column_length):
            current_row = start_row + i  # Adjust to paste into the correct row
            source_ws[f'{target_column}{current_row}'] = cell_value
            print(f"Pasting '{cell_value}' into {target_column}{current_row}")

    except Exception as e:
        print(f"Unexpected error: {str(e)}")

# Function to convert .xls to .xlsx and transfer data to a backup of ASN copy
def convert_xls_data(uploaded_file, dest_file):
    xls_book = xlrd.open_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    source_ws = source_wb.active
    xls_sheet = xls_book.sheet_by_index(0)

    # Mapping uploaded cells to copy cells
    data_map = {
        (17, 1): 'E4',   # 'B18' -> (18, 2) name
        (17, 3): 'E5',   # 'D18' -> (18, 4) number
        (17, 4): 'E6',   # 'E18' -> (18, 5) add 1
        (17, 9): 'E7',   # 'J18' -> (18, 10) city
        (17, 10): 'E8',  # 'K18' -> (18, 11) State
        (17, 11): 'E9',  # 'L18' -> (18, 12) Zip
        (9, 2): 'E14'    # 'C10' -> (10, 3) delivery date
    }

    for (row, col), copy_cell in data_map.items():
        value = xls_sheet.cell_value(row, col)
        source_ws[copy_cell] = value

    # Manually assign values for cells that don’t come from the xls_sheet
    source_ws['B12'] = "FEDG"
    source_ws['B13'] = "Fedex"
    
    # Set current date in E11 and highlight it
    current_date = datetime.datetime.now().strftime("%m/%d/%Y")
    source_ws['E11'] = current_date
    
    # Highlight cell E11 with yellow background
    from openpyxl.styles import PatternFill
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    source_ws['E11'].fill = yellow_fill

    # Debugging: Print the total number of rows and columns in the uploaded sheet
    total_rows = xls_sheet.nrows
    total_cols = xls_sheet.ncols
    print(f"Total rows: {total_rows}, Total columns: {total_cols}")

    # Ensure you don't go out of bounds
    if total_rows < 23:
        print(f"Error: Not enough rows to start processing from row 23.")
        return

    # Create carton-based lines instead of product-based lines
    carton_lines_count = create_carton_based_lines(xls_sheet, source_ws, start_row=23, target_start_row=19)
    
    print(f"Total carton lines created: {carton_lines_count}")
    
    # Additional copy and paste operations with helper functions
    oneToMany(xls_sheet, source_ws, row=3, col=2, target_column='B', start_row=19, column_length=carton_lines_count)  # PO
    oneToMany(xls_sheet, source_ws, row=3, col=7, target_column='C', start_row=19, column_length=carton_lines_count)  # PO Date

    typedValue(source_ws, static_value="NA", target_column='D', start_row=19, column_length=carton_lines_count)

    # Set total cartons in B15 and "C" in E13
    source_ws['B15'] = carton_lines_count
    source_ws['E13'] = "C"

    format_cells_as_text(source_ws)
    format_cells_as_text(source_ws)
    align_cells_left(source_ws)
    align_cells_left(source_ws)

    # Save the updated file
    try:
        source_wb.save(dest_file)
        print(f"File saved successfully as {dest_file}.")
    except Exception as e:
        print(f"Error saving file: {str(e)}")


def process_MurdochsASN(file_path):
    """Main function to process Murdochs ASN files."""
    current_date = datetime.datetime.now().strftime("%m.%d.%Y")
    # Determine if the file is XLSX or XLS and extract the PO number
    if file_path.endswith('.xlsx'):
        uploaded_wb = load_workbook(file_path)
        po_number = uploaded_wb.active['C4'].value
    elif file_path.endswith('.xls'):
        xls_book = xlrd.open_workbook(file_path)
        po_number = xls_book.sheet_by_index(0).cell_value(3, 2)

    # Define the backup file path in FINISHED_FOLDER with a Murdochs subfolder
    backup_file = os.path.join(FINISHED_FOLDER, f"Murdochs/Murdochs 856 ASN PO {po_number} {current_date}.xlsx")
    
    # Ensure the directory exists
    os.makedirs(os.path.dirname(backup_file), exist_ok=True)

    # Perform the copy or conversion based on file type
    if file_path.endswith('.xlsx'):
        copy_xlsx_data(file_path, backup_file)
    elif file_path.endswith('.xls'):
        convert_xls_data(file_path, backup_file)

    return backup_file, po_number

def process_murdochs_asn_with_labels(order_file_path, edi_file_path=None):
    """
    Complete two-step ASN processing workflow.
    
    Step 1: Process order file to create ASN with carton-based lines
    Step 2: (Optional) Process EDI file to add carton labels
    
    Args:
        order_file_path: Path to the original order file (.xls or .xlsx)
        edi_file_path: (Optional) Path to EDI file with carton labels
    
    Returns:
        tuple: (final_asn_file_path, po_number)
    
    Example Usage:
        # Step 1 only: Create ASN from order
        asn_file, po = process_murdochs_asn_with_labels("order.xlsx")
        
        # Step 1 + 2: Create ASN and add carton labels
        final_asn, po = process_murdochs_asn_with_labels("order.xlsx", "edi_labels.xlsx")
    """
    print("=== ENTERING process_murdochs_asn_with_labels ===")
    print(f"Order file path: {order_file_path}")
    print(f"EDI file path: {edi_file_path}")
    
    print("=== STEP 1: Processing Order File ===")
    # Step 1: Create ASN from order file
    asn_file_path, po_number = process_MurdochsASN(order_file_path)
    print(f"Step 1 completed. ASN file: {asn_file_path}")
    
    if edi_file_path:
        print("\n=== STEP 2: Processing EDI Carton Labels ===")
        # Step 2: Add carton labels from EDI file
        final_asn_path = process_carton_labels_edi(edi_file_path, asn_file_path)
        print(f"Step 2 completed. Final ASN file: {final_asn_path}")
        return final_asn_path, po_number
    else:
        print("\nNo EDI file provided. ASN created without carton labels.")
        return asn_file_path, po_number

def process_carton_labels_edi(edi_file_path, existing_asn_file_path):
    """
    Second step: Process EDI file with 20-digit carton labels and populate them into existing ASN file.
    
    Args:
        edi_file_path: Path to the EDI file containing carton labels
        existing_asn_file_path: Path to the existing ASN file to update
    
    Returns:
        str: Path to the updated ASN file
    """
    print(f"Processing EDI carton labels file: {edi_file_path}")
    print(f"Updating existing ASN file: {existing_asn_file_path}")
    
    # Load the existing ASN file
    asn_wb = load_workbook(existing_asn_file_path)
    asn_ws = asn_wb.active
    
    # Determine EDI file type and load it
    print(f"EDI file path: {edi_file_path}")
    print(f"EDI file extension: {edi_file_path.split('.')[-1]}")
    
    if edi_file_path.endswith('.xlsx'):
        print("Processing as XLSX file...")
        edi_wb = load_workbook(edi_file_path)
        edi_ws = edi_wb.active
        carton_labels = extract_carton_labels_from_xlsx(edi_ws)
    elif edi_file_path.endswith('.xls'):
        print("Processing as XLS file...")
        try:
            edi_book = xlrd.open_workbook(edi_file_path)
            print(f"Successfully opened XLS file with {len(edi_book.sheet_names())} sheets")
            print(f"Sheet names: {edi_book.sheet_names()}")
            
            # Try to find carton labels in all sheets
            carton_labels = []
            for sheet_idx, sheet_name in enumerate(edi_book.sheet_names()):
                print(f"\n=== CHECKING SHEET {sheet_idx + 1}: {sheet_name} ===")
                edi_sheet = edi_book.sheet_by_index(sheet_idx)
                sheet_labels = extract_carton_labels_from_xls(edi_sheet)
                if sheet_labels:
                    print(f"Found {len(sheet_labels)} carton labels in sheet '{sheet_name}'")
                    carton_labels.extend(sheet_labels)
                else:
                    print(f"No carton labels found in sheet '{sheet_name}'")
            
            if not carton_labels:
                print(f"\n=== NO CARTON LABELS FOUND IN ANY SHEET ===")
                print("Available sheets:")
                for i, name in enumerate(edi_book.sheet_names()):
                    sheet = edi_book.sheet_by_index(i)
                    print(f"  Sheet {i+1}: '{name}' - {sheet.nrows} rows, {sheet.ncols} columns")
                    
        except Exception as e:
            print(f"Error opening XLS file: {e}")
            raise
    else:
        raise ValueError("EDI file must be .xlsx or .xls format")
    
    # Populate carton labels into the ASN file
    populate_carton_labels_in_asn(asn_ws, carton_labels)
    
    # Save the updated ASN file
    updated_file_path = existing_asn_file_path.replace('.xlsx', '_with_labels.xlsx')
    asn_wb.save(updated_file_path)
    
    print(f"Updated ASN file saved as: {updated_file_path}")
    return updated_file_path

def extract_carton_labels_from_xlsx(edi_ws):
    """
    Extract carton labels from Column E and product names from Column I of XLSX EDI file, starting from row 19.
    Returns a list of dictionaries with carton label data and associated product names.
    """
    carton_labels = []
    
    print(f"\n=== EXTRACTING CARTON LABELS AND PRODUCT NAMES FROM XLSX EDI FILE ===")
    
    # Look for data starting from row 19 (where data typically begins in ASN files)
    row = 19
    print(f"Starting to look for carton labels in Column E and product names in Column I from row {row}")
    
    while edi_ws[f'E{row}'].value is not None:
        try:
            # Extract carton label from column E
            carton_label = edi_ws[f'E{row}'].value  # 20-digit carton label
            
            # Extract product name from column I
            product_name = edi_ws[f'I{row}'].value  # Product name/SKU
            
            print(f"Row {row}: Carton Label = '{carton_label}', Product Name = '{product_name}'")
            
            if carton_label:
                carton_label_str = str(carton_label).strip()
                # Check if it's a 20-digit number (all digits)
                if len(carton_label_str) == 20 and carton_label_str.isdigit():
                    carton_labels.append({
                        'carton_label': carton_label_str,
                        'product_name': str(product_name).strip() if product_name else '',
                        'row_number': row
                    })
                    print(f"  ✓ Added carton label: {carton_label_str} with product: '{product_name}'")
                else:
                    print(f"  ✗ Skipped (not a 20-digit number: {carton_label_str})")
            else:
                print(f"  ✗ Skipped (empty cell)")
            
            row += 1
            
        except Exception as e:
            print(f"Error processing EDI row {row}: {e}")
            row += 1
            continue
    
    print(f"Extracted {len(carton_labels)} carton labels with product names from EDI file")
    return carton_labels

def extract_carton_labels_from_xls(edi_sheet):
    """
    Extract carton labels from Column E and product names from Column I of XLS EDI file, starting from row 19.
    Returns a list of dictionaries with carton label data and associated product names.
    """
    carton_labels = []
    
    print(f"\n=== EXTRACTING CARTON LABELS AND PRODUCT NAMES FROM XLS EDI FILE ===")
    print(f"Total rows in EDI sheet: {edi_sheet.nrows}")
    print(f"Total columns in EDI sheet: {edi_sheet.ncols}")
    
    # Look for carton labels in Column E (index 4) and product names in Column I (index 8) starting from row 19
    print(f"\n=== LOOKING FOR CARTON LABELS IN COLUMN E AND PRODUCT NAMES IN COLUMN I (starting from row 19) ===")
    
    for row_idx in range(18, edi_sheet.nrows):  # Start from row 19 (index 18)
        try:
            # Extract carton label from column E (index 4)
            carton_label = edi_sheet.cell_value(row_idx, 4)  # Column E = index 4
            
            # Extract product name from column I (index 8)
            product_name = edi_sheet.cell_value(row_idx, 8)  # Column I = index 8
            
            print(f"Row {row_idx + 1}: Carton Label = '{carton_label}' (type: {type(carton_label)}), Product Name = '{product_name}'")
            
            # Handle different data types that xlrd might return
            if carton_label:
                carton_label_str = str(carton_label).strip()
                print(f"  Converted to string: '{carton_label_str}' (length: {len(carton_label_str)})")
                
                # Check if it's a 20-digit number (all digits)
                if len(carton_label_str) == 20 and carton_label_str.isdigit():
                    carton_labels.append({
                        'carton_label': carton_label_str,
                        'product_name': str(product_name).strip() if product_name else '',
                        'row_number': row_idx + 1
                    })
                    print(f"  ✓ Added carton label: {carton_label_str} with product: '{product_name}'")
                else:
                    print(f"  ✗ Skipped (not a 20-digit number: {carton_label_str})")
            else:
                print(f"  ✗ Skipped (empty cell)")
                
        except Exception as e:
            print(f"Error processing EDI row {row_idx + 1}: {e}")
            continue
    
    print(f"\n=== SUMMARY ===")
    print(f"Found {len(carton_labels)} carton labels with product names in Column E")
    
    # If we found carton labels, show them
    if carton_labels:
        print("Carton labels found:")
        for i, label in enumerate(carton_labels):
            print(f"  {i+1}. {label['carton_label']} (Row {label['row_number']}) with product: '{label['product_name']}'")
    else:
        print("No 20-digit carton labels found in Column E!")
    
    return carton_labels

def populate_carton_labels_in_asn(asn_ws, carton_labels):
    """
    Populate carton labels into the ASN file using intelligent product name matching.
    This ensures each 20-digit label is matched to the correct product regardless of order.
    """
    print(f"\n=== INTELLIGENT CARTON LABEL MATCHING ===")
    print(f"Available carton labels: {len(carton_labels)}")
    
    # Extract product names from the ASN file (column H, starting from row 19)
    asn_product_names = extract_product_names_from_asn(asn_ws)
    print(f"Found {len(asn_product_names)} products in ASN file")
    
    # Extract product names from the EDI file to create the mapping
    edi_product_names = extract_product_names_from_edi_labels(carton_labels)
    print(f"Found {len(edi_product_names)} products in EDI file")
    
    # Create intelligent mapping between EDI products and ASN products
    product_mapping = create_product_name_mapping(edi_product_names, asn_product_names)
    print(f"Created mapping for {len(product_mapping)} products")
    
    # Populate carton labels based on the mapping
    populated_count = populate_labels_using_mapping(asn_ws, carton_labels, product_mapping)
    
    print(f"\n=== SUMMARY ===")
    print(f"Successfully matched and populated {populated_count} carton labels using intelligent product name matching")

def extract_product_names_from_asn(asn_ws):
    """
    Extract product names from column I of the ASN file, starting from row 19.
    Returns a list of dictionaries with row number and product name.
    """
    product_names = []
    row = 19
    
    print(f"\n=== EXTRACTING PRODUCT NAMES FROM ASN FILE ===")
    print(f"Looking for product names in column I starting from row {row}")
    
    while asn_ws[f'I{row}'].value is not None:
        product_name = asn_ws[f'I{row}'].value
        if product_name and str(product_name).strip():
            product_names.append({
                'row': row,
                'name': str(product_name).strip(),
                'original_name': product_name
            })
            print(f"Row {row}: Product name = '{product_name}'")
        row += 1
    
    print(f"Extracted {len(product_names)} product names from ASN file")
    return product_names

def extract_product_names_from_edi_labels(carton_labels):
    """
    Extract product names from the EDI file based on carton label positions.
    Now that we're extracting both carton labels and product names, we can use this data directly.
    Returns a list of dictionaries with carton label and associated product name.
    """
    print(f"\n=== EXTRACTING PRODUCT NAMES FROM EDI LABELS ===")
    
    edi_products = []
    
    for label_data in carton_labels:
        if 'product_name' in label_data and label_data['product_name']:
            edi_products.append({
                'carton_label': label_data['carton_label'],
                'name': label_data['product_name'],
                'row_number': label_data['row_number']
            })
            print(f"Product: '{label_data['product_name']}' with carton label: {label_data['carton_label']}")
    
    print(f"Extracted {len(edi_products)} product names from EDI labels")
    return edi_products

def create_product_name_mapping(edi_products, asn_products):
    """
    Create intelligent mapping between EDI products and ASN products based on name similarity.
    Uses fuzzy string matching to find the best matches.
    """
    print(f"\n=== CREATING PRODUCT NAME MAPPING ===")
    
    if not edi_products or not asn_products:
        print("Warning: Missing product data for mapping")
        return {}
    
    mapping = {}
    used_asn_products = set()
    
    # For each EDI product, find the best matching ASN product
    for edi_product in edi_products:
        best_match = None
        best_score = 0
        
        for asn_product in asn_products:
            if asn_product['row'] in used_asn_products:
                continue  # This ASN product is already matched
                
            # Calculate similarity score using simple string matching
            score = calculate_name_similarity(edi_product['name'], asn_product['name'])
            
            if score > best_score:
                best_score = score
                best_match = asn_product
        
        if best_match and best_score > 0.5:  # Threshold for acceptable match
            mapping[edi_product['carton_label']] = {
                'asn_row': best_match['row'],
                'edi_name': edi_product['name'],
                'asn_name': best_match['name'],
                'similarity_score': best_score
            }
            used_asn_products.add(best_match['row'])
            print(f"Matched: '{edi_product['name']}' -> '{best_match['name']}' (score: {best_score:.2f})")
        else:
            print(f"Warning: No good match found for '{edi_product['name']}' (best score: {best_score:.2f})")
    
    return mapping

def calculate_name_similarity(name1, name2):
    """
    Calculate similarity between two product names.
    Returns a score between 0 and 1, where 1 is exact match.
    """
    if not name1 or not name2:
        return 0
    
    # Convert to lowercase and remove common separators
    name1_clean = name1.lower().replace('-', ' ').replace('_', ' ').strip()
    name2_clean = name2.lower().replace('-', ' ').replace('_', ' ').strip()
    
    # Exact match
    if name1_clean == name2_clean:
        return 1.0
    
    # Check if one contains the other
    if name1_clean in name2_clean or name2_clean in name1_clean:
        return 0.8
    
    # Check for common words
    words1 = set(name1_clean.split())
    words2 = set(name2_clean.split())
    
    if words1 and words2:
        common_words = words1.intersection(words2)
        total_words = words1.union(words2)
        if total_words:
            return len(common_words) / len(total_words)
    
    return 0.0

def populate_labels_using_mapping(asn_ws, carton_labels, product_mapping):
    """
    Populate carton labels into the ASN file using the product mapping.
    """
    print(f"\n=== POPULATING LABELS USING MAPPING ===")
    
    populated_count = 0
    unmapped_labels = []
    
    for carton_label_data in carton_labels:
        carton_label = carton_label_data['carton_label']
        
        if carton_label in product_mapping:
            mapping = product_mapping[carton_label]
            asn_row = mapping['asn_row']
            
            # Populate the carton label in column E of the ASN
            asn_ws[f'E{asn_row}'] = carton_label
            print(f"Row {asn_row}: Mapped carton label {carton_label} to product '{mapping['asn_name']}' (similarity: {mapping['similarity_score']:.2f})")
            populated_count += 1
        else:
            unmapped_labels.append(carton_label)
            print(f"Warning: No mapping found for carton label {carton_label}")
    
    print(f"\n=== MAPPING SUMMARY ===")
    print(f"Successfully populated {populated_count} carton labels using intelligent matching")
    
    if unmapped_labels:
        print(f"Warning: {len(unmapped_labels)} carton labels could not be mapped:")
        for label in unmapped_labels:
            print(f"  - {label}")
        print("These labels may not have had good product name matches or the products may not exist in the ASN file.")
    
    # Report any unused ASN rows
    total_asn_products = len([p for p in extract_product_names_from_asn(asn_ws) if p['name']])
    if populated_count < total_asn_products:
        remaining = total_asn_products - populated_count
        print(f"Note: {remaining} ASN product rows were not populated with carton labels")
        print("This may be normal if there are more products in the ASN than carton labels in the EDI file.")
    
    return populated_count

def extract_company_and_po_from_edi(edi_file_path):
    """
    Extract company name and PO number from EDI file.
    For Murdochs: Company name is in A1, PO number is in B19.
    Returns a tuple of (company_name, po_number) or (None, None) if not found.
    """
    try:
        if edi_file_path.endswith('.xlsx'):
            return extract_company_and_po_from_xlsx(edi_file_path)
        elif edi_file_path.endswith('.xls'):
            return extract_company_and_po_from_xls(edi_file_path)
        else:
            print(f"Unsupported file format: {edi_file_path}")
            return None, None
    except Exception as e:
        print(f"Error extracting company and PO from EDI file: {e}")
        return None, None

def extract_company_and_po_from_xlsx(edi_file_path):
    """
    Extract company name and PO number from XLSX EDI file.
    For Murdochs: Company name is in A1, PO number is in B19.
    """
    try:
        edi_wb = load_workbook(edi_file_path)
        edi_ws = edi_wb.active
        
        # Extract company name from A1
        company_name = edi_ws['A1'].value
        print(f"Extracted company name from A1: '{company_name}'")
        
        # Extract PO number from B19
        po_number = edi_ws['B19'].value
        print(f"Extracted PO number from B19: '{po_number}'")
        
        return company_name, po_number
        
    except Exception as e:
        print(f"Error extracting company and PO from XLSX EDI file: {e}")
        return None, None

def extract_company_and_po_from_xls(edi_file_path):
    """
    Extract company name and PO number from XLS EDI file.
    For Murdochs: Company name is in A1 (row 0, col 0), PO number is in B19 (row 18, col 1).
    """
    try:
        edi_wb = xlrd.open_workbook(edi_file_path)
        edi_sheet = edi_wb.sheet_by_index(0)
        
        # Extract company name from A1 (row 0, col 0)
        company_name = edi_sheet.cell_value(0, 0)
        print(f"Extracted company name from A1: '{company_name}'")
        
        # Extract PO number from B19 (row 18, col 1)
        po_number = edi_sheet.cell_value(18, 1)
        print(f"Extracted PO number from B19: '{po_number}'")
        
        return company_name, po_number
        
    except Exception as e:
        print(f"Error extracting company and PO from XLS EDI file: {e}")
        return None, None

def find_matching_asn_file(company_name, po_number):
    """
    Find an existing ASN file that matches the company name and PO number.
    Returns the full path to the matching ASN file or None if not found.
    """
    try:
        print(f"Searching for ASN file with company: '{company_name}', PO: '{po_number}'")
        
        # Normalize the company name for comparison
        company_normalized = company_name.strip().upper() if company_name else ""
        po_normalized = str(po_number).strip() if po_number else ""
        
        print(f"Normalized search terms - Company: '{company_normalized}', PO: '{po_normalized}'")
        
        # Search through all ASN files in the Finished folder
        for root, dirs, filenames in os.walk(FINISHED_FOLDER):
            for filename in filenames:
                if filename.endswith('.xlsx') and 'ASN' in filename:
                    file_path = os.path.join(root, filename)
                    
                    # Check if this is a Murdochs ASN file
                    if 'Murdochs' in filename or 'murdochs' in filename.lower():
                        print(f"Checking Murdochs ASN file: {filename}")
                        
                        # Extract PO from the ASN file
                        try:
                            asn_wb = load_workbook(file_path)
                            asn_ws = asn_wb.active
                            
                            # Get PO from B19 in the ASN file
                            asn_po = asn_ws['B19'].value
                            asn_po_normalized = str(asn_po).strip() if asn_po else ""
                            
                            print(f"  ASN file PO: '{asn_po_normalized}'")
                            
                            # Check if PO numbers match
                            if asn_po_normalized == po_normalized:
                                print(f"  ✓ MATCH FOUND: {file_path}")
                                return file_path
                            else:
                                print(f"  ✗ PO mismatch")
                                
                        except Exception as e:
                            print(f"  Error reading ASN file {filename}: {e}")
                            continue
        
        print(f"No matching ASN file found for company '{company_name}' and PO '{po_number}'")
        return None
        
    except Exception as e:
        print(f"Error finding matching ASN file: {e}")
        return None

def process_edi_with_auto_match(edi_file_path):
    """
    Process EDI file by automatically finding the matching ASN file.
    Returns the path to the updated ASN file or None if no match found.
    """
    try:
        print(f"\n=== AUTO-MATCHING EDI TO ASN ===")
        print(f"EDI file: {edi_file_path}")
        
        # Extract company and PO from EDI file
        company_name, po_number = extract_company_and_po_from_edi(edi_file_path)
        
        if not company_name or not po_number:
            print("Could not extract company name or PO number from EDI file")
            return None
        
        # Find matching ASN file
        matching_asn_path = find_matching_asn_file(company_name, po_number)
        
        if not matching_asn_path:
            print("No matching ASN file found")
            return None
        
        # Process the carton labels
        print(f"Processing carton labels for matched ASN: {matching_asn_path}")
        updated_asn_path = process_carton_labels_edi(edi_file_path, matching_asn_path)
        
        return updated_asn_path
        
    except Exception as e:
        print(f"Error in auto-match process: {e}")
        return None
