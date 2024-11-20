from openpyxl import load_workbook
import xlrd
import os
from datetime import datetime
from DigitHelper import process_20_digit_labels, process_single_label_with_description, get_filled_rows_count
from ExcelHelpers import FINISHED_FOLDER

def process_Thrive20(asn_file_path, label_file_path, finished_folder):
    from openpyxl import load_workbook
    import xlrd
    import os
    import datetime

    current_date = datetime.datetime.now().strftime("%m.%d.%Y")

    # Determine the PO number
    if asn_file_path.endswith('.xlsx'):
        uploaded_wb = load_workbook(asn_file_path)
        po_number = uploaded_wb.active['C4'].value
    elif asn_file_path.endswith('.xls'):
        xls_book = xlrd.open_workbook(asn_file_path)
        po_number = xls_book.sheet_by_index(0).cell_value(3, 2)
    else:
        raise ValueError("Unsupported file format for ASN file.")

    # Generate output file path
    output_file_path = os.path.join(
        finished_folder, f"Thrive/Thrive 856 ASN PO {po_number} {current_date}.xlsx"
    )
    
    # Process the files (add your Thrive-specific logic here)
    # Example: Copy the label from the label file to the ASN file
    xls_book = xlrd.open_workbook(label_file_path)
    label_sheet = xls_book.sheet_by_index(0)
    label_value = label_sheet.cell_value(18, 3)  # Get 20-digit label from D19 (zero-based index)

    # Load the ASN file and update it
    asn_wb = load_workbook(asn_file_path)
    asn_ws = asn_wb.active
    column_length = get_filled_rows_count(asn_ws, start_row=19, start_col=1)  # Use helper function

    # Update the column with the label value
    for row in range(19, 19 + column_length):
        asn_ws[f"D{row}"] = label_value

    # Save the updated ASN file
    os.makedirs(os.path.dirname(output_file_path), exist_ok=True)
    asn_wb.save(output_file_path)
    print(f"Updated ASN file saved to: {output_file_path}")
    
    return output_file_path, po_number

