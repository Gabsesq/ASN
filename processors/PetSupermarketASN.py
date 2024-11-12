from openpyxl import load_workbook
import xlrd
import os
import datetime
from ExcelHelpers import (
    resource_path, FINISHED_FOLDER, format_cells_as_text, align_cells_left
)

# Define source files and destination copies for Pet Supermarket
source_asn_xlsx = resource_path("assets/Pet Supermarket/Blank Pet Supermarket ASN.xlsx")

# Function to copy data from uploaded .xlsx file to specific cells in the ASN .xlsx backup
def copy_xlsx_data(uploaded_file, dest_file):
    uploaded_wb = load_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    uploaded_ws = uploaded_wb.active
    source_ws = source_wb.active

    # Copy data from column A (starting at row 16) in the uploaded file to column A (starting at row 19) in the output file
    row = 16
    data_to_copy = []
    while uploaded_ws[f'A{row}'].value is not None:
        data_to_copy.append(uploaded_ws[f'A{row}'].value)
        row += 1

    # Paste data into column A of the output file starting from row 19
    for i, value in enumerate(data_to_copy):
        source_ws[f'A{19 + i}'] = value

    # Get the value from C4 in the upload file and paste into column B (starting at row 19)
    value_from_upload = uploaded_ws['C4'].value
    if value_from_upload is not None:
        for i in range(19, 19 + len(data_to_copy)):
            source_ws[f'B{i}'] = value_from_upload

    # Copy H4 from upload file into C19 down to the number of rows copied above
    value_from_H4 = uploaded_ws['H4'].value
    if value_from_H4 is not None:
        for i in range(19, 19 + len(data_to_copy)):
            source_ws[f'C{i}'] = value_from_H4

    # Copy data from F16 down in the upload file into E19 down in the output file
    row = 16
    for i in range(19, 19 + len(data_to_copy)):
        source_ws[f'E{i}'] = uploaded_ws[f'F{row}'].value
        row += 1

    # Copy data from G16 down into F19 down in the output file
    row = 16
    for i in range(19, 19 + len(data_to_copy)):
        source_ws[f'F{i}'] = uploaded_ws[f'G{row}'].value
        row += 1

    # Copy data from H16 down into G19 down in the output file
    row = 16
    for i in range(19, 19 + len(data_to_copy)):
        source_ws[f'G{i}'] = uploaded_ws[f'H{row}'].value
        row += 1

    # Copy data from B16 down into H19 down in the output file
    row = 16
    for i in range(19, 19 + len(data_to_copy)):
        source_ws[f'H{i}'] = uploaded_ws[f'B{row}'].value
        row += 1

    # Copy data from C16 down into I19 down in the output file
    row = 16
    for i in range(19, 19 + len(data_to_copy)):
        source_ws[f'I{i}'] = uploaded_ws[f'C{row}'].value
        row += 1

    # Copy data from E16 down into J19 down in the output file
    row = 16
    for i in range(19, 19 + len(data_to_copy)):
        source_ws[f'J{i}'] = uploaded_ws[f'E{row}'].value
        row += 1

    # Copy data from I16 down into K19 down in the output file
    row = 16
    for i in range(19, 19 + len(data_to_copy)):
        source_ws[f'K{i}'] = uploaded_ws[f'I{row}'].value
        row += 1

    format_cells_as_text(source_ws)
    align_cells_left(source_ws)

    # Save the updated copy
    source_wb.save(dest_file)

# Function to convert .xls to .xlsx and transfer data to a backup of ASN copy
def convert_xls_data(uploaded_file, dest_file):
    xls_book = xlrd.open_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    source_ws = source_wb.active
    xls_sheet = xls_book.sheet_by_index(0)
    

    # Copy the value from C7 in the upload file into B12 in the output file
    delivery_date = xls_sheet.cell_value(6, 2)  # Row 7 (index 6), Column C (index 2)
    source_ws['B12'] = delivery_date

    # Copy data from column A (starting at row 16) in the uploaded file to column A (starting at row 19) in the output file
    row = 16
    data_to_copy = []
    while True:
        try:
            value = xls_sheet.cell_value(row - 1, 0)  # Column A is index 0
            if value:
                data_to_copy.append(value)
                row += 1
            else:
                break
        except IndexError:
            break

    # Paste data into column A of the output file starting from row 19
    for i, value in enumerate(data_to_copy):
        source_ws[f'A{19 + i}'] = value

    # Get the value from C4 in the upload file and paste into column B (starting at row 19)
    value_from_upload_C4 = xls_sheet.cell_value(3, 2)
    if value_from_upload_C4 is not None:
        for i in range(19, 19 + len(data_to_copy)):
            source_ws[f'B{i}'] = value_from_upload_C4

    # Copy H4 from upload file into C19 down to the number of rows copied above
    value_from_H4 = xls_sheet.cell_value(3, 7)
    if value_from_H4 is not None:
        for i in range(19, 19 + len(data_to_copy)):
            source_ws[f'C{i}'] = value_from_H4

    # Copy data from F16 down in the upload file into E19 down in the output file
    row = 16
    for i in range(19, 19 + len(data_to_copy)):
        source_ws[f'E{i}'] = xls_sheet.cell_value(row - 1, 5)
        row += 1

    # Copy data from G16 down into F19 down in the output file
    row = 16
    for i in range(19, 19 + len(data_to_copy)):
        source_ws[f'F{i}'] = xls_sheet.cell_value(row - 1, 6)
        row += 1

    # Copy data from H16 down into G19 down in the output file
    row = 16
    for i in range(19, 19 + len(data_to_copy)):
        source_ws[f'G{i}'] = xls_sheet.cell_value(row - 1, 7)
        row += 1

    # Copy data from B16 down into H19 down in the output file
    row = 16
    for i in range(19, 19 + len(data_to_copy)):
        source_ws[f'H{i}'] = xls_sheet.cell_value(row - 1, 1)
        row += 1

    # Copy data from C16 down into I19 down in the output file
    row = 16
    for i in range(19, 19 + len(data_to_copy)):
        source_ws[f'I{i}'] = xls_sheet.cell_value(row - 1, 2)
        row += 1

    # Copy data from E16 down into J19 down in the output file
    row = 16
    for i in range(19, 19 + len(data_to_copy)):
        source_ws[f'J{i}'] = xls_sheet.cell_value(row - 1, 4)
        row += 1

    # Copy data from I16 down into K19 down in the output file
    row = 16
    for i in range(19, 19 + len(data_to_copy)):
        source_ws[f'K{i}'] = xls_sheet.cell_value(row - 1, 8)
        row += 1

    format_cells_as_text(source_ws)
    align_cells_left(source_ws)
    align_cells_left(source_ws)
    # Save the updated copy
    source_wb.save(dest_file)

def process_PetSupermarketASN(file_path):
    """Main function to process Pet Supermarket ASN files."""
    current_date = datetime.datetime.now().strftime("%m.%d.%Y")

    # Determine if the file is XLSX or XLS and extract the PO number
    if file_path.endswith('.xlsx'):
        uploaded_wb = load_workbook(file_path)
        po_number = uploaded_wb.active['C4'].value
    elif file_path.endswith('.xls'):
        xls_book = xlrd.open_workbook(file_path)
        po_number = xls_book.sheet_by_index(0).cell_value(3, 2)

    # Define the backup file path in FINISHED_FOLDER with a PetSupermarket subfolder
    backup_file = os.path.join(FINISHED_FOLDER, f"PetSupermarket/Pet Supermarket ASN PO {po_number} {current_date}.xlsx")
    
    # Ensure the directory exists
    os.makedirs(os.path.dirname(backup_file), exist_ok=True)

    # Perform the copy or conversion based on file type
    if file_path.endswith('.xlsx'):
        copy_xlsx_data(file_path, backup_file)
    elif file_path.endswith('.xls'):
        convert_xls_data(file_path, backup_file)

    return backup_file, po_number
