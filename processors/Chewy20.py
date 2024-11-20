from DigitHelper import process_20_digit_labels, process_single_label_with_description
from datetime import datetime
from openpyxl import load_workbook

def process_Chewy20(asn_file_path, carton_label_file_path):
    config = {
        "start_row_asn": 20,  # Starting row for ASN file processing
        "start_row_label": 20,  # Starting row for label file processing
        "label_col": 4,  # Column D for carton labels
        "upc_col": 5,  # Column E for UPCs
        "vendor_part_col": 6,  # Column F for Vendor Part
        "sku_col": 7,  # Column G for SKU
        "description_col": 11,  # Column K for Description
    }

    # Check the number of filled rows in ASN file
    asn_wb = load_workbook(asn_file_path)
    asn_ws = asn_wb.active
    filled_rows = 0
    for row in asn_ws.iter_rows(min_row=config["start_row_asn"], max_col=1, values_only=True):
        if row[0]:  # Check if column A is not empty
            filled_rows += 1

    print(f"Filled rows from A{config['start_row_asn']} down: {filled_rows}")

    if filled_rows > 10:
        output_path = process_single_label_with_description(
            asn_file_path,
            carton_label_file_path,
            config,
            output_file_name="Updated_Chewy_ASN_SingleLabel.xlsx",
            mixed_description="Mixed"
        )
    else:
        output_path, _ = process_20_digit_labels(
            asn_file_path,
            carton_label_file_path,
            config,
            output_file_name="Updated_Chewy_ASN.xlsx"
        )

    # Add the current date to cell E11
    current_date = datetime.now().strftime("%m/%d/%Y")
    asn_ws = load_workbook(output_path).active
    asn_ws["E11"] = current_date
    print(f"Pasting current date '{current_date}' into E11")
    asn_ws.parent.save(output_path)

    return output_path, None
