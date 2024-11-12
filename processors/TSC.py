from openpyxl import load_workbook
import xlrd
import datetime
import os
import sys
import tempfile
from ExcelHelpers import (
    QTY_total, resource_path, oneToMany, manyToMany, get_current_date, extract_po_number, format_cells_as_text, align_cells_left, get_column_length, FINISHED_FOLDER
)



# Define source file for TSC
source_asn_xlsx = resource_path("assets/TSC/Blank TSC ASN.xlsx")
print(f"Resolved path for source_asn_xlsx: {source_asn_xlsx}")

# Check if the file exists
if not os.path.exists(source_asn_xlsx):
    print("File does not exist at:", source_asn_xlsx)
else:
    print("File found at:", source_asn_xlsx)


# Function to copy data from uploaded .xlsx file to the ASN .xlsx backup
def copy_xlsx_data(uploaded_file, dest_file):
    try:
        uploaded_wb = load_workbook(uploaded_file)
        source_wb = load_workbook(source_asn_xlsx)
        uploaded_ws = uploaded_wb.active
        source_ws = source_wb.active

        format_cells_as_text(source_ws)
        align_cells_left(source_ws)

        # Mapping uploaded cells to copy cells
        data_map = {
            'B14': 'E3', 'D14': 'E4', 'E14': 'E5',
            'J14': 'E6', 'K14': 'E7', 'L14': 'E8',
            'C9': 'B11'
        }

        # Transfer data with debugging
        for upload_cell, copy_cell in data_map.items():
            try:
                value = uploaded_ws[upload_cell].value
                source_ws[copy_cell] = value
                print(f"Copied value '{value}' from {upload_cell} to {copy_cell}")
            except Exception as e:
                print(f"Error copying from {upload_cell} to {copy_cell}: {str(e)}")

        # Calculate dynamic column length starting from A18
        column_length = get_column_length(uploaded_ws, start_row=18)

        # Copy value from 'C4' into 'B17' and down using oneToMany
        oneToMany(uploaded_ws, source_ws, 3, 2, 'B', 17, column_length)

        format_cells_as_text(source_ws)
        align_cells_left(source_ws)
        align_cells_left(source_ws)
        source_wb.save(dest_file)
        print(f"Saved file successfully as {dest_file}")

    except Exception as e:
        print(f"Error in copy_xlsx_data: {str(e)}")

# Function to convert and copy data from .xls files
def convert_xls_data(uploaded_file, dest_file):
    try:
        xls_book = xlrd.open_workbook(uploaded_file)
        source_wb = load_workbook(source_asn_xlsx)
        source_ws = source_wb.active
        xls_sheet = xls_book.sheet_by_index(0)

        format_cells_as_text(source_ws)
        align_cells_left(source_ws)

        # Mapping uploaded cells to copy cells with error handling
        data_map = {
            (13, 1): 'E3', (13, 3): 'E4', (13, 4): 'E5',
            (13, 9): 'E6', (13, 10): 'E7', (13, 11): 'E8',
            (8, 2): 'B11', (6,2) : 'E11'
        }
        for (row, col), copy_cell in data_map.items():
            try:
                value = xls_sheet.cell_value(row, col)
                source_ws[copy_cell] = value
                print(f"Copied value '{value}' from ({row}, {col}) to {copy_cell}")
            except IndexError as e:
                print(f"IndexError: ({row}, {col}) is out of range: {str(e)}")
            except Exception as e:
                print(f"Unexpected error copying ({row}, {col}) to {copy_cell}: {str(e)}")

        # Calculate dynamic column length starting from A17
        column_length = get_column_length(xls_sheet, start_row=17)

        # Perform many-to-many copy operations with debugging
        manyToMany(xls_sheet, source_ws, 18, 0, 'A', 17, column_length)  # Item No
        oneToMany(xls_sheet, source_ws, 3, 2, 'B', 17, column_length-1)    # Copy 'C4' to 'B17'
        manyToMany(xls_sheet, source_ws, 18, 4, 'D', 17, column_length)  # UPS
        manyToMany(xls_sheet, source_ws, 18, 5, 'E', 17, column_length)  # Buyer Part
        manyToMany(xls_sheet, source_ws, 18, 6, 'F', 17, column_length)  # Vendor Part
        manyToMany(xls_sheet, source_ws, 18, 1, 'G', 17, column_length)  # QTY
        manyToMany(xls_sheet, source_ws, 18, 2, 'H', 17, column_length)  # UOM
        manyToMany(xls_sheet, source_ws, 18, 8, 'I', 17, column_length)  # Description

        # Sum the QTY values from column B (index 1) and place the total in E13
        qty_total = QTY_total(xls_sheet, 18 , 1)
        source_ws['E13'] = qty_total
        print(f"Total QTY placed in E13: {qty_total}")


        source_wb.save(dest_file)
        print(f"Saved file successfully as {dest_file}")

    except Exception as e:
        print(f"Error in convert_xls_data: {str(e)}")


def process_TSC(file_path):
    """Main function to process TSC files."""
    try:
        current_date = get_current_date()
        
        # Define output in a persistent FINISHED_FOLDER directory
        po_number = extract_po_number(file_path, is_xlsx=file_path.endswith('.xlsx'))
        backup_file = os.path.join(FINISHED_FOLDER, f"TSC/Tractor Supply ASN {po_number} {current_date}.xlsx")
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(backup_file), exist_ok=True)

        if file_path.endswith('.xlsx'):
            copy_xlsx_data(file_path, backup_file)
        elif file_path.endswith('.xls'):
            convert_xls_data(file_path, backup_file)
        
        # Return backup_file path for further processing
        return backup_file, po_number

    except Exception as e:
        print(f"Error in process_TSC: {str(e)}")
        return None  # Ensure None is returned if an error occurs


