from openpyxl import load_workbook
import xlrd
import datetime
import os
from ExcelHelpers import (
    resource_path, FINISHED_FOLDER, format_cells_as_text, align_cells_left, manyToMany, oneToMany, typedValue
)
from upc_counts import counts  # Import the UPC counts dictionary

# Define source files and destination copies for Chewy
source_asn_xlsx = resource_path("assets/Murdochs/Blank Murdochs UCC128 Label Request.xlsx")

# Function to copy data from uploaded .xlsx file to specific cells in the ASN .xlsx backup
def copy_xlsx_data(uploaded_file, dest_file):
    uploaded_wb = load_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    uploaded_ws = uploaded_wb.active
    source_ws = source_wb.active
    format_cells_as_text(source_ws)
    align_cells_left(source_ws)
    # Mapping uploaded cells to copy cells
    data_map = {
        'B18': 'E3', 'D18': 'E4', 'E18': 'E5',
        'J18': 'E6', 'K18': 'E7', 'L18': 'E8',
        'C10': 'E14'
    }

    # Transfer data from the uploaded file to the backup copy
    for upload_cell, copy_cell in data_map.items():
        source_ws[copy_cell] = uploaded_ws[upload_cell].value

    # Track numbers from A21 and below, copy to A20 in the copy
    row = 21
    data_to_copy = []
    while uploaded_ws[f'A{row}'].value is not None:
        data_to_copy.append(uploaded_ws[f'A{row}'].value)
        row += 1

    for i, value in enumerate(data_to_copy):
        source_ws[f'A{20 + i}'] = value

    count = len(data_to_copy)
    value_from_upload = uploaded_ws['C4'].value

    if value_from_upload is not None:
        for i in range(20, 20 + count):
            source_ws[f'B{i}'] = value_from_upload

    source_wb.save(dest_file)

# Function to convert .xls to .xlsx and transfer data to a backup of ASN copy
def convert_xls_data(uploaded_file, dest_file):
    xls_book = xlrd.open_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    source_ws = source_wb.active
    xls_sheet = xls_book.sheet_by_index(0)
    # Mapping uploaded cells to copy cells
    data_map = {
        (17, 1): 'F3',   # 'B18' -> (18, 2) name
        (17, 4): 'F4',   # 'E18' -> (18, 5) add 1
        (17, 5): 'F5',   # 'E18' -> (18, 5) add 2
        (17, 9): 'F6',  # 'J18' -> (18, 10) city
        (17, 10): 'F7',  # 'K18' -> (18, 11) State
        (17, 11): 'F8',  # 'L18' -> (18, 12) Zip
    }
    

    for (row, col), copy_cell in data_map.items():
        value = xls_sheet.cell_value(row, col)
        source_ws[copy_cell] = value

      # Debugging: Print the total number of rows and columns in the uploaded sheet
        total_rows = xls_sheet.nrows
        total_cols = xls_sheet.ncols
        print(f"Total rows: {total_rows}, Total columns: {total_cols}")

        # Ensure you don't go out of bounds
        if total_rows < 23:
            print(f"Error: Not enough rows to start processing from row 23.")
            return

        # Count total number of data rows (looking for non-empty values in column A starting from row 23)
        data_rows = 0
        for row_idx in range(22, xls_sheet.nrows):  # Start from index 22 (row 23)
            try:
                value = xls_sheet.cell_value(row_idx, 0)  # Column A (index 0)
                if value != '':  # Check for any non-empty value
                    data_rows += 1
            except IndexError:
                break
        
        print(f"Total data rows found: {data_rows}")

        # Function to calculate cases based on UPC and quantity
        def calculate_cases(upc, qty):
            if upc in counts:
                items_per_case = counts[upc]
                return qty // items_per_case  # Integer division to get whole cases
            print(f"Warning: UPC {upc} not found in counts dictionary")
            return 0

        try:
            if data_rows > 0:
                # Copy existing data
                manyToMany(xls_sheet, source_ws, 23, 6, 'F', 14, data_rows)  # Vendor Part number
                oneToMany(xls_sheet, source_ws, row=3, col=2, target_column='A', start_row=14, column_length=data_rows)  # PO
                oneToMany(xls_sheet, source_ws, row=17, col=11, target_column='B', start_row=14, column_length=data_rows)  # Zip
                typedValue(source_ws, static_value="Fed Ex", target_column='C', start_row=14, column_length=data_rows)

                # Calculate and fill in case quantities
                for i in range(data_rows):
                    row_idx = i + 22  # Starting from row 23 (index 22)
                    qty = int(xls_sheet.cell_value(row_idx, 1))  # Column B (QTY)
                    upc = str(int(xls_sheet.cell_value(row_idx, 5)))  # Column F (UPC)
                    cases = calculate_cases(upc, qty)
                    source_ws[f'H{14 + i}'] = cases  # Write cases to column H
                    print(f"Row {row_idx + 1}: UPC {upc}, QTY {qty}, Cases {cases}")

            else:
                print("No data found in column A starting from row 23.")
        except Exception as e:
            print(f"Error during copy operations: {str(e)}")

        format_cells_as_text(source_ws)
        align_cells_left(source_ws)
        align_cells_left(source_ws)
        # Save the updated file
        try:
            source_wb.save(dest_file)
            print(f"File saved successfully as {dest_file}.")
        except Exception as e:
            print(f"Error saving file: {str(e)}")

def process_MurdochsLabel(file_path):
    """Main function to process Murdochs UCC128 Label Request files."""
    current_date = datetime.datetime.now().strftime("%m.%d.%Y")

    # Determine if the file is XLSX or XLS and extract the PO number
    if file_path.endswith('.xlsx'):
        uploaded_wb = load_workbook(file_path)
        po_number = uploaded_wb.active['C4'].value
    elif file_path.endswith('.xls'):
        xls_book = xlrd.open_workbook(file_path)
        po_number = xls_book.sheet_by_index(0).cell_value(3, 2)

    # Define the backup file path in FINISHED_FOLDER with a Murdochs subfolder
    backup_file = os.path.join(FINISHED_FOLDER, f"Murdochs/Murdochs UCC128 Label Request {po_number} {current_date}.xlsx")
    
    # Ensure the directory exists
    os.makedirs(os.path.dirname(backup_file), exist_ok=True)

    # Perform the copy or conversion based on file type
    if file_path.endswith('.xlsx'):
        copy_xlsx_data(file_path, backup_file)
    elif file_path.endswith('.xls'):
        convert_xls_data(file_path, backup_file)

    return backup_file, po_number
