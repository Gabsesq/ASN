from openpyxl import load_workbook
import xlrd
import datetime
import os
from ExcelHelpers import (
    resource_path, FINISHED_FOLDER, format_cells_as_text, align_cells_left, manyToMany, oneToMany, typedValue, QTY_total, get_column_length
)

# Define source files and destination copies for Chewy
source_asn_xlsx = resource_path("assets/Thrive Market/Blank Thrive Market UCC128 Label Request 7.19.24.xlsx")

# Function to copy data from uploaded .xlsx file to specific cells in the ASN .xlsx backup
def copy_xlsx_data(uploaded_file, dest_file):
    uploaded_wb = load_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    uploaded_ws = uploaded_wb.active
    source_ws = source_wb.active
    format_cells_as_text(source_ws)
    align_cells_left(source_ws)
    # Mapping uploaded cells to copy cells
    data_map = {
        'B18': 'E3', 'D18': 'E4', 'E18': 'E5',
        'J18': 'E6', 'K18': 'E7', 'L18': 'E8',
        'C10': 'E14'
    }

    # Transfer data from the uploaded file to the backup copy
    for upload_cell, copy_cell in data_map.items():
        source_ws[copy_cell] = uploaded_ws[upload_cell].value

    # Track numbers from A21 and below, copy to A20 in the copy
    row = 21
    data_to_copy = []
    while uploaded_ws[f'A{row}'].value is not None:
        data_to_copy.append(uploaded_ws[f'A{row}'].value)
        row += 1

    for i, value in enumerate(data_to_copy):
        source_ws[f'A{20 + i}'] = value

    count = len(data_to_copy)
    value_from_upload = uploaded_ws['C4'].value

    if value_from_upload is not None:
        for i in range(20, 20 + count):
            source_ws[f'B{i}'] = value_from_upload

    source_wb.save(dest_file)

# Function to convert .xls to .xlsx and transfer data to a backup of ASN copy
def convert_xls_data(uploaded_file, dest_file):
    xls_book = xlrd.open_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    source_ws = source_wb.active
    xls_sheet = xls_book.sheet_by_index(0)


    # Mapping uploaded cells to copy cells
    data_map = {
        (12, 1): 'F3',   # Name
        (12, 4): 'F4',   # Address 1
        (12, 7): 'F5',   # Address 2
        (12, 9): 'F6',   # City
        (12, 10): 'F7',  # State
        (12, 11): 'F8',  # Zip
    }

    for (row, col), copy_cell in data_map.items():
        value = xls_sheet.cell_value(row, col)
        source_ws[copy_cell] = value

    # Calculate total quantity
    total_qty = QTY_total(xls_sheet, start_row=17, qty_column=1)

    # Logic for total_qty > 10
    if total_qty <= 10:
        print(f"Total Quantity: {total_qty}. Performing extended operations...")

        column_length = get_column_length(xls_sheet, start_row=17)
        # oneToMany(xls_sheet, source_ws, row, col, target_column, start_row, column_length):
        # manyToMany(xls_sheet, source_ws, start_row, start_col, dest_col, dest_start_row, column_length):
        
        if column_length > 0:
            oneToMany(xls_sheet, source_ws, 3, 2, 'A', 14, column_length)  # PO number
            oneToMany(xls_sheet, source_ws, 12, 11, 'B', 14, column_length) # zipcode
            typedValue(source_ws, "Fedex", 'C', 14, column_length) # Carrier name
            manyToMany(xls_sheet, source_ws, 17, 5, 'E', 14, column_length) # UPC
            typedValue(source_ws, "1", 'F', 14, column_length) # Carton QTY, always 1
            manyToMany(xls_sheet, source_ws, 17, 4, 'G', 14, column_length) # Description
            manyToMany(xls_sheet, source_ws, 17, 1, 'H', 14, column_length) # Labels



    # Logic for total_qty > 10
    else:
        print(f"Total Quantity: {total_qty}. Performing basic operations...")
        # Basic data mapping for smaller quantities
        source_ws["E14"] = "mixed"   # UPC
        data_map2 = {
            (3, 2): 'A14',   # PO
            (11, 11): 'B14',   # Zip
            "SAIA": 'C14',    # Carrier
            total_qty: 'F14',  #QTY
            "mixed": 'G14',   # Description
            "1": 'H14'        # Labels
        }

        # Loop through data_map2 and set values accordingly
        for key, copy_cell in data_map2.items():
            if isinstance(key, tuple):
                # Key is a tuple (row, col) - use cell value from xls_sheet
                row, col = key
                value = xls_sheet.cell_value(row, col)
            elif isinstance(key, (int, float)):
                # Key is an integer or float - likely total_qty or other numbers
                value = key
            else:
                # Key is a string - directly use the string value
                value = key
            
            # Assign the value to the corresponding cell in source_ws
            source_ws[copy_cell] = value


    # Save the updated file
    try:
        source_wb.save(dest_file)
        print(f"File saved successfully as {dest_file}.")
    except Exception as e:
        print(f"Error saving file: {str(e)}")



def process_ThriveLabel(file_path):
    current_date = datetime.datetime.now().strftime("%m.%d.%Y")

    # Determine if the file is XLSX or XLS
    if file_path.endswith('.xlsx'):
        uploaded_wb = load_workbook(file_path)
        po_number = uploaded_wb.active['C4'].value

    elif file_path.endswith('.xls'):
        xls_book = xlrd.open_workbook(file_path)
        po_number = xls_book.sheet_by_index(0).cell_value(3, 2)

    # Define the backup file path in the FINISHED_FOLDER using resource_path
    backup_file = os.path.join(FINISHED_FOLDER, f"Thrive/Thrive Market UCC128 Label Request {po_number} {current_date}.xlsx")

     # Ensure the Thrive directory exists in the FINISHED_FOLDER
    os.makedirs(os.path.dirname(backup_file), exist_ok=True)

    # Perform the copy or conversion based on file type
    if file_path.endswith('.xlsx'):
        copy_xlsx_data(file_path, backup_file)
    elif file_path.endswith('.xls'):
        convert_xls_data(file_path, backup_file)

    return backup_file, po_number