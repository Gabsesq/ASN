from openpyxl import load_workbook
import xlrd
import os
from ExcelHelpers import (
    resource_path, FINISHED_FOLDER, format_cells_as_text, align_cells_left, manyToMany, oneToMany,
    typedValue, QTY_total, get_column_length, create_folder, extract_po_number, get_current_date, generate_rows
)

# Define source ASN template file path
source_asn_xlsx = resource_path("assets/Thrive Market/Thrive Market 856 Master Template.xlsx")



def copy_xlsx_data(uploaded_file, dest_file):
    """Copy data from an uploaded .xlsx file to the ASN backup."""
    uploaded_wb = load_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    uploaded_ws = uploaded_wb.active
    source_ws = source_wb.active

    format_cells_as_text(source_ws)
    align_cells_left(source_ws)


    # Copy data dynamically using helper functions
    manyToMany(uploaded_ws, source_ws, 17, 0, 'A', 19, get_column_length(uploaded_ws, 17))
    oneToMany(uploaded_ws, source_ws, 3, 2, 'B', 19, get_column_length(uploaded_ws, 17))
    oneToMany(uploaded_ws, source_ws, 3, 7, 'C', 19, get_column_length(uploaded_ws, 17))

    # Use a static value for a specific column
    typedValue(source_ws, "Fed Ex", 'D', 19, get_column_length(uploaded_ws, 17))

    # Calculate total quantity and place it in 'F14'
    total_qty = QTY_total(uploaded_ws, 17, 1)
    source_ws['F14'] = total_qty

    # Save the updated workbook
    source_wb.save(dest_file)
    print(f"File saved successfully as {dest_file}.")

def convert_xls_data(uploaded_file, dest_file):
    """Convert and copy data from .xls to .xlsx backup."""
    xls_book = xlrd.open_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    source_ws = source_wb.active
    xls_sheet = xls_book.sheet_by_index(0)

    format_cells_as_text(source_ws)
    align_cells_left(source_ws)

    # Define your data maps
    data_map = {
        (11, 1): 'E4', (11, 3): 'E5', (11, 4):'E6',
        (11, 9): 'E7', (11, 10): 'E8', (11, 11): 'E9',
    }


    for (row, col), copy_cell in data_map.items():
        value = xls_sheet.cell_value(row, col)
        source_ws[copy_cell] = value

    
    # Starting row in output sheet where data should be copied
    output_row = 19  # Destination row start
    item_start_row = 17  # Source row start for items in xls
    total_lines = 0

    print("Processing rows and duplicating based on QTY...")

    # Loop through items in the source .xls sheet
    while True:
        try:
            # Read QTY and other necessary fields
            qty = int(xls_sheet.cell_value(item_start_row - 1, 1))  # Column B for QTY
            po_number = xls_sheet.cell_value(3, 2)  # PO Number (row 3, column 2)
            po_date = xls_sheet.cell_value(3, 7)  # PO Date (row 3, column 7)
            vendor_part = xls_sheet.cell_value(item_start_row - 1, 6)  # Column F for Vendor Part
            upc = xls_sheet.cell_value(item_start_row - 1, 5)  # Column E for UPC
            pack = xls_sheet.cell_value(item_start_row - 1, 7)  # Column I for Pack
            description = xls_sheet.cell_value(item_start_row - 1, 4)  # Column H for Description
            uom = xls_sheet.cell_value(item_start_row - 1, 2)  # Column C for UOM

            # Duplicate rows based on QTY
            for _ in range(qty):
                source_ws[f'A{output_row}'] = output_row - 18  # Line number
                source_ws[f'B{output_row}'] = po_number
                source_ws[f'C{output_row}'] = po_date
                source_ws[f'F{output_row}'] = upc
                source_ws[f'E{output_row}'] = vendor_part
                source_ws[f'G{output_row}'] = "1"
                source_ws[f'H{output_row}'] = uom
                source_ws[f'I{output_row}'] = pack
                source_ws[f'J{output_row}'] = description
                output_row += 1  # Move to the next row

            total_lines += qty  # Increment total lines
            item_start_row += 1  # Move to the next source row

            # Stop if no more data is found in the QTY column
            if not xls_sheet.cell_value(item_start_row - 1, 6):
                break

        except IndexError:
            # Reached the end of the source sheet
            break

    # Write total QTY to a specific cell
    print(f"Total QTY processed: {total_lines}")
    source_ws['F15'] = total_lines

    # Apply formatting
    format_cells_as_text(source_ws)
    align_cells_left(source_ws)

    # Save the updated workbook
    source_wb.save(dest_file)
    print(f"File saved successfully as {dest_file}.")

def process_ThriveASN(file_path):
    """Main function to process Thrive ASN files."""
    current_date = get_current_date()
    po_number = extract_po_number(file_path, file_path.endswith('.xlsx'))

    # Define the destination backup file path
    backup_file = os.path.join(FINISHED_FOLDER, f"Thrive/Thrive ASN PO {po_number} {current_date}.xlsx")
    
    # Ensure the Thrive folder exists in the Finished directory
    os.makedirs(os.path.dirname(backup_file), exist_ok=True)
    print(f"Resolved backup file path: {backup_file}")  # Debug line to confirm path

    # Process based on file type
    if file_path.endswith('.xlsx'):
        copy_xlsx_data(file_path, backup_file)
    elif file_path.endswith('.xls'):
        convert_xls_data(file_path, backup_file)

    return backup_file, po_number
