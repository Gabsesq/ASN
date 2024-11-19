from openpyxl import load_workbook
import xlrd
import datetime
import os
from ExcelHelpers import (
    oneToMany, resource_path, FINISHED_FOLDER, format_cells_as_text, align_cells_left, QTY_total
)

# Define source files and destination copies for Chewy
source_asn_xlsx = resource_path("assets/Chewy/Chewy 856 ASN - Copy.xlsx")

# Function to copy data from uploaded .xlsx file to specific cells in the ASN .xlsx backup
def copy_xlsx_data(uploaded_file, dest_file):
    uploaded_wb = load_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    uploaded_ws = uploaded_wb.active
    source_ws = source_wb.active


    
    # Mapping uploaded cells to copy cells
    data_map = {
        'B16': 'E3', 'D16': 'E4', 'E16': 'E5',
        'J16': 'E6', 'K16': 'E7', 'L16': 'E8',
        'C8': 'E14'
    }

    # Transfer data from the uploaded file to the backup copy
    for upload_cell, copy_cell in data_map.items():
        source_ws[copy_cell] = uploaded_ws[upload_cell].value

    # Track numbers from A21 and below, copy to A20 in the copy
    row = 21
    data_to_copy = []
    while uploaded_ws[f'A{row}'].value is not None:
        data_to_copy.append(uploaded_ws[f'A{row}'].value)
        row += 1

    for i, value in enumerate(data_to_copy):
        source_ws[f'A{20 + i}'] = value

    count = len(data_to_copy)
    value_from_upload = uploaded_ws['C4'].value

    if value_from_upload is not None:
        for i in range(20, 20 + count):
            source_ws[f'B{i}'] = value_from_upload

    source_wb.save(dest_file)

# Function to convert .xls to .xlsx and transfer data to a backup of ASN copy
def convert_xls_data(uploaded_file, dest_file):
    xls_book = xlrd.open_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    source_ws = source_wb.active
    xls_sheet = xls_book.sheet_by_index(0)

    # Mapping uploaded cells to copy cells
    data_map = {
        (15, 1): 'E4', (15, 3): 'E5', (15, 4): 'E6',
        (15, 9): 'E7', (15, 10): 'E8', (15, 11): 'E9',
        (7, 2): 'E15'
    }

    for (row, col), copy_cell in data_map.items():
        value = xls_sheet.cell_value(row, col)
        source_ws[copy_cell] = value

    # Starting row in output sheet where data should be copied
    output_row = 20
    total_lines = 0

    # Loop through each item in the upload sheet
    item_start_row = 21  # Row index where items start in the upload sheet
    while True:
        try:
            # Read QTY and other fields from the upload sheet
            qty = int(xls_sheet.cell_value(item_start_row - 1, 1))  # Column B for QTY (0-based index)
            description = xls_sheet.cell_value(item_start_row - 1, 4)  # Column E for Description
            vendor_part = xls_sheet.cell_value(item_start_row - 1, 6)  # Column G for Vendor Part #
            upc14 = xls_sheet.cell_value(item_start_row - 1, 7)  # Column H for UPC14
            sku = xls_sheet.cell_value(item_start_row - 1, 8)  # Column I for SKU
            total_lines += qty


            # Repeat each field `qty` times in the output sheet
            for _ in range(qty):
                source_ws[f'A{output_row}'] = output_row - 19  # Item No count (1, 2, 3, ...)
                source_ws[f'B{output_row}'] = qty  # QTY
                source_ws[f'L{output_row}'] = description
                source_ws[f'E{output_row}'] = upc14  # Placeholder for UPC if needed
                source_ws[f'F{output_row}'] = vendor_part
                source_ws[f'H{output_row}'] = "1"  # Assuming UOM is always "CA", adjust as needed
                source_ws[f'G{output_row}'] = sku
                output_row += 1  # Move to the next row in the output sheet

                                # Conditional logic based on QTY value
            if total_lines > 10:
                # Define data map for cases where QTY > 10
                data_map = {
                    'B12': 'SAIA',  # Adding word values directly
                    'B15': 'SAIA',
                    'E13': 'P',
                }
            else:
                # Define data map for cases where QTY <= 10
                data_map = {
                    'B12': 'FEDG',
                    'B15': 'Fedex',
                    'E15': total_lines,
                    'E13': 'C',
                }

            # Apply data mapping based on the condition
            for src_cell, value in data_map.items():
                # Instead of fetching value from xls_sheet, we use the literal string value
                source_ws[src_cell] = value


            item_start_row += 1  # Move to the next item row in the upload sheet

            # Stop if there's no more data in the QTY column
            if not xls_sheet.cell_value(item_start_row - 1, 1):
                break

        except IndexError:
            # Reached the end of the data
            break
    


    # Use oneToMany for additional fields if required (example for C4, PO)
    oneToMany(xls_sheet, source_ws, row=3, col=2, target_column='B', start_row=20, column_length=output_row - 20) #PO
    oneToMany(xls_sheet, source_ws, row=3, col=7, target_column='C', start_row=20, column_length=output_row - 20) #PO Date
    oneToMany(xls_sheet, source_ws, row=20, col=2, target_column='I', start_row=20, column_length=output_row - 20)

    format_cells_as_text(source_ws)
    align_cells_left(source_ws)
    align_cells_left(source_ws)

    # Save the updated copy
    source_wb.save(dest_file)

def process_ChewyASN(file_path):
    """Main function to process Chewy ASN files."""
    current_date = datetime.datetime.now().strftime("%m.%d.%Y")

    # Determine if the file is XLSX or XLS and extract the PO number
    if file_path.endswith('.xlsx'):
        uploaded_wb = load_workbook(file_path)
        po_number = uploaded_wb.active['C4'].value
    elif file_path.endswith('.xls'):
        xls_book = xlrd.open_workbook(file_path)
        po_number = xls_book.sheet_by_index(0).cell_value(3, 2)

    # Define the backup file path in FINISHED_FOLDER with a Chewy subfolder
    backup_file = os.path.join(FINISHED_FOLDER, f"Chewy/Chewy 856 ASN PO {po_number} {current_date}.xlsx")
    
    # Ensure the directory exists
    os.makedirs(os.path.dirname(backup_file), exist_ok=True)

    # Perform the copy or conversion based on file type
    if file_path.endswith('.xlsx'):
        copy_xlsx_data(file_path, backup_file)
    elif file_path.endswith('.xls'):
        convert_xls_data(file_path, backup_file)

    return backup_file, po_number
