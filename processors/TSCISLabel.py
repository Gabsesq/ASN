from openpyxl import load_workbook
import xlrd
import datetime
import os
import sys
import tempfile
from ExcelHelpers import (
    QTY_total, resource_path, oneToMany, manyToMany, get_current_date, extract_po_number, format_cells_as_text, align_cells_left, get_column_length, FINISHED_FOLDER
)
from upc_counts import counts  # Add this import



# Define source file for TSC IS - Update the path to match your actual template file
source_asn_xls = resource_path("assets/TSCIS/Master Template TSC IS UCC128 Label Request.xlsx")  # Changed template name

# Add debug print to help troubleshoot
print(f"Looking for Label template at: {source_asn_xls}")
if not os.path.exists(source_asn_xls):
    print(f"WARNING: Label template file not found at {source_asn_xls}")
else:
    print(f"Label template file found at {source_asn_xls}")


# Function to copy data from uploaded .xlsx file to the ASN .xlsx backup
def copy_xlsx_data(uploaded_file, dest_file):
    try:
        # For xlsx files, we'll convert the template first
        output_wb = load_workbook(source_asn_xls.replace('.xls', '.xlsx'))
        uploaded_wb = load_workbook(uploaded_file)
        
        output_ws = output_wb.active
        uploaded_ws = uploaded_wb.active

        format_cells_as_text(output_ws)
        align_cells_left(output_ws)

        # Mapping uploaded cells to copy cells
        data_map = {
            'B14': 'E3', 'D14': 'E4', 'E14': 'E5',
            'J14': 'E6', 'K14': 'E7', 'L14': 'E8',
            'C9': 'B11'
        }

        # Transfer static data
        for upload_cell, copy_cell in data_map.items():
            try:
                value = uploaded_ws[upload_cell].value
                output_ws[copy_cell] = value
                print(f"Copied value '{value}' from {upload_cell} to {copy_cell}")
            except Exception as e:
                print(f"Error copying from {upload_cell} to {copy_cell}: {str(e)}")

        # Calculate total labels needed based on UPC counts
        total_labels = 0
        row = 18  # Start at row 18
        
        while uploaded_ws[f'B{row}'].value:  # While we have QTY values
            try:
                upc = str(int(uploaded_ws[f'E{row}'].value))  # Get UPC from column E
                qty = int(uploaded_ws[f'B{row}'].value)       # Get QTY from column B
                
                if upc in counts:
                    items_per_case = counts[upc]
                    labels_needed = qty // items_per_case  # Integer division
                    total_labels += labels_needed
                    print(f"Row {row}: UPC {upc}, QTY {qty}, Items/Case {items_per_case}, Labels {labels_needed}")
                else:
                    print(f"Warning: UPC {upc} not found in counts dictionary")
                
                row += 1
                
            except Exception as e:
                print(f"Error processing row {row}: {str(e)}")
                break

        # Write total labels to C14
        output_ws['C14'] = total_labels
        print(f"Total labels needed: {total_labels}")

        output_wb.save(dest_file)
        return True

    except Exception as e:
        print(f"Error in copy_xlsx_data: {str(e)}")
        return False

# Function to convert and copy data from .xls files
def convert_xls_data(uploaded_file, dest_file):
    try:
        # Create a new workbook from the template
        source_wb = load_workbook(source_asn_xls)
        source_ws = source_wb.active
        
        # Read the uploaded file using xlrd
        xls_book = xlrd.open_workbook(uploaded_file)
        xls_sheet = xls_book.sheet_by_index(0)

        format_cells_as_text(source_ws)
        align_cells_left(source_ws)

        # Mapping uploaded cells to copy cells with error handling
        data_map = {
            (13, 1): 'F3', (13, 4): 'F4', (13,7): 'F5',
            (13, 9): 'F6', (13, 10): 'F7', (13, 11): 'F8',
            (3, 2): 'A14', (13, 3):'B14'
        }
        
        # Copy the static data first
        for (row, col), copy_cell in data_map.items():
            try:
                value = xls_sheet.cell_value(row, col)
                source_ws[copy_cell] = value
                print(f"Copied value '{value}' from ({row}, {col}) to {copy_cell}")
            except Exception as e:
                print(f"Error copying ({row}, {col}) to {copy_cell}: {str(e)}")

        # Calculate total labels needed based on UPC counts
        total_labels = 0
        row = 17  # Start at row 18 (17 in 0-based index)
        
        while True:
            try:
                # Get UPC and QTY from the current row
                upc = str(int(xls_sheet.cell_value(row, 4)))  # Column E (4 in 0-based)
                qty = int(xls_sheet.cell_value(row, 1))      # Column B (1 in 0-based)
                
                if upc in counts:
                    items_per_case = counts[upc]
                    labels_needed = qty // items_per_case  # Integer division
                    total_labels += labels_needed
                    print(f"Row {row+1}: UPC {upc}, QTY {qty}, Items/Case {items_per_case}, Labels {labels_needed}")
                else:
                    print(f"Warning: UPC {upc} not found in counts dictionary")
                
                row += 1  # Move to next row
                
                # Check if we've reached the end of the data
                if not xls_sheet.cell_value(row, 1):  # No more QTY values
                    break
                    
            except IndexError:
                # Reached end of sheet
                break
            except Exception as e:
                print(f"Error processing row {row+1}: {str(e)}")
                break

        # Write total labels to C14
        source_ws['C14'] = total_labels
        print(f"Total labels needed: {total_labels}")

        source_wb.save(dest_file)
        print(f"Saved file successfully as {dest_file}")
        return True

    except Exception as e:
        print(f"Error in convert_xls_data: {str(e)}")
        return False


def process_TSCISLabel(file_path):
    """Main function to process TSCIS Label files."""
    try:
        current_date = get_current_date()
        po_number = extract_po_number(file_path, is_xlsx=file_path.endswith('.xlsx'))
        
        # Use a different filename for the label file
        backup_file = os.path.join(FINISHED_FOLDER, 'TSCIS', f"Tractor Supply IS Label {po_number} {current_date}.xlsx")
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(backup_file), exist_ok=True)

        success = False
        if file_path.endswith('.xlsx'):
            success = copy_xlsx_data(file_path, backup_file)
        elif file_path.endswith('.xls'):
            success = convert_xls_data(file_path, backup_file)
        
        if not success:
            raise Exception("Failed to process file")
            
        if not os.path.exists(backup_file):
            raise Exception(f"Output file was not created at {backup_file}")
            
        return backup_file, po_number

    except Exception as e:
        print(f"Error in process_TSCISLabel: {str(e)}")
        raise
