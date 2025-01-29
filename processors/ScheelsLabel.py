from openpyxl import load_workbook
import xlrd
import datetime
import os
from ExcelHelpers import (
    get_column_length, resource_path, FINISHED_FOLDER, format_cells_as_text, align_cells_left, manyToMany, oneToMany, typedValue
)

# Define source files and destination copies for Chewy
source_asn_xlsx = resource_path("assets\Scheels\Blank Scheels UCC128 Label Request.xlsx")

# Function to copy data from uploaded .xlsx file to specific cells in the ASN .xlsx backup
def copy_xlsx_data(uploaded_file, dest_file):
    uploaded_wb = load_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    uploaded_ws = uploaded_wb.active
    source_ws = source_wb.active
    format_cells_as_text(source_ws)
    align_cells_left(source_ws)
    # Mapping uploaded cells to copy cells
    data_map = {
        'B18': 'E3', 'D18': 'E4', 'E18': 'E5',
        'J18': 'E6', 'K18': 'E7', 'L18': 'E8',
        'C10': 'E14'
    }

    # Transfer data from the uploaded file to the backup copy
    for upload_cell, copy_cell in data_map.items():
        source_ws[copy_cell] = uploaded_ws[upload_cell].value

    # Track numbers from A21 and below, copy to A20 in the copy
    row = 21
    data_to_copy = []
    while uploaded_ws[f'A{row}'].value is not None:
        data_to_copy.append(uploaded_ws[f'A{row}'].value)
        row += 1

    for i, value in enumerate(data_to_copy):
        source_ws[f'A{20 + i}'] = value

    count = len(data_to_copy)
    value_from_upload = uploaded_ws['C4'].value

    if value_from_upload is not None:
        for i in range(20, 20 + count):
            source_ws[f'B{i}'] = value_from_upload

    source_wb.save(dest_file)

# Function to convert .xls to .xlsx and transfer data to a backup of ASN copy
def convert_xls_data(uploaded_file, dest_file):
    xls_book = xlrd.open_workbook(uploaded_file)
    source_wb = load_workbook(source_asn_xlsx)
    source_ws = source_wb.active
    xls_sheet = xls_book.sheet_by_index(0)
    format_cells_as_text(source_ws)
    align_cells_left(source_ws)
    
    # Mapping uploaded cells to copy cells
    data_map = {
        (12, 1): 'F3',   # 'B18' -> (18, 2) name
        (12, 4): 'F4',   # 'E18' -> (18, 5) add 1
        (12, 5): 'F5',   # 'E18' -> (18, 5) add 2
        (12, 9): 'F6',  # 'J18' -> (18, 10) city
        (12, 10): 'F7',  # 'K18' -> (18, 11) State
        (12, 11): 'F8',  # 'L18' -> (18, 12) Zip
    }
    
    # Process the data map first
    for (row, col), copy_cell in data_map.items():
        value = xls_sheet.cell_value(row, col)
        source_ws[copy_cell] = value

    # Dynamic column length calculation with boundary check
    start_row = 17
    column_length = get_column_length(xls_sheet, start_row)
    print("Final Column Length:", column_length)  # Confirm calculated column length

    # Now process all the repeated data outside the loop
    oneToMany(xls_sheet=xls_sheet, source_ws=source_ws, row=3, col=2, target_column='A', start_row=14, column_length=column_length) #PO
    oneToMany(xls_sheet=xls_sheet, source_ws=source_ws, row=12, col=11, target_column='B', start_row=14, column_length=column_length) #Zip Code
    
    typedValue(source_ws=source_ws, static_value="UPS", target_column='C', start_row=14, column_length=column_length) #UPS
    typedValue(source_ws=source_ws, static_value="NA", target_column='F', start_row=14, column_length=column_length) #Special Number
    typedValue(source_ws=source_ws, static_value="NA", target_column='H', start_row=14, column_length=column_length) #Mark For

    manyToMany(xls_sheet=xls_sheet, source_ws=source_ws, start_row=17, start_col=5, dest_col='G', dest_start_row=14, column_length=column_length)  # UPC  

    format_cells_as_text(source_ws)
    align_cells_left(source_ws)
    align_cells_left(source_ws)

    # Save the updated file
    try:
        source_wb.save(dest_file)
        print(f"File saved successfully as {dest_file}.")
    except Exception as e:
        print(f"Error saving file: {str(e)}")
    print("column length ", column_length)



def process_ScheelsLabel(file_path):
    """Main function to process Scheels UCC128 Label Request files."""
    current_date = datetime.datetime.now().strftime("%m.%d.%Y")

    # Determine if the file is XLSX or XLS and extract the PO number
    if file_path.endswith('.xlsx'):
        uploaded_wb = load_workbook(file_path)
        po_number = uploaded_wb.active['C4'].value
    elif file_path.endswith('.xls'):
        xls_book = xlrd.open_workbook(file_path)
        po_number = xls_book.sheet_by_index(0).cell_value(3, 2)

    # Define the backup file path in FINISHED_FOLDER with a Scheels subfolder
    backup_file = os.path.join(FINISHED_FOLDER, f"Scheels/Scheels UCC128 Label Request {po_number} {current_date}.xlsx")
    
    # Ensure the directory exists
    os.makedirs(os.path.dirname(backup_file), exist_ok=True)

    # Perform the copy or conversion based on file type
    if file_path.endswith('.xlsx'):
        copy_xlsx_data(file_path, backup_file)
    elif file_path.endswith('.xls'):
        convert_xls_data(file_path, backup_file)

    return backup_file, po_number

