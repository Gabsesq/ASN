from openpyxl import load_workbook
import xlrd
import datetime
import os
from ExcelHelpers import (
    resource_path, FINISHED_FOLDER, format_cells_as_text, align_cells_left, manyToMany, oneToMany
)

# Define source files and destination copies for Chewy
source_label_xlsx = resource_path("assets/Chewy/Chewy UCC128 Label Request - Copy.xlsx")

# Function to copy data from uploaded .xlsx file to specific cells in the ASN .xlsx backup
def copy_xlsx_data(uploaded_file, dest_file):
    uploaded_wb = load_workbook(uploaded_file)
    source_wb = load_workbook(source_label_xlsx)
    uploaded_ws = uploaded_wb.active
    source_ws = source_wb.active

    format_cells_as_text(source_ws)
    align_cells_left(source_ws)

    # Mapping uploaded cells to copy cells
    data_map = {
        'B16': 'F3', 'E16': 'F4', 'H16': 'F5',
        'J16': 'F6', 'K16': 'F7', 'L16': 'F8'
    }

   # Transfer data from the uploaded file to the backup copy
    for upload_cell, copy_cell in data_map.items():
        source_ws[copy_cell] = uploaded_ws[upload_cell].value

    # Track numbers from A21 and below, copy to A20 in the copy
    row = 21
    data_to_copy = []
    while uploaded_ws[f'A{row}'].value is not None:
        data_to_copy.append(uploaded_ws[f'A{row}'].value)
        row += 1

    for i, value in enumerate(data_to_copy):
        source_ws[f'A{20 + i}'] = value

    count = len(data_to_copy)
    value_from_upload = uploaded_ws['C4'].value

    if value_from_upload is not None:
        for i in range(20, 20 + count):
            source_ws[f'B{i}'] = value_from_upload

    source_wb.save(dest_file)

# Function to convert .xls to .xlsx and transfer data to a backup of ASN copy
def convert_xls_data(uploaded_file, dest_file):
    xls_book = xlrd.open_workbook(uploaded_file)
    source_wb = load_workbook(source_label_xlsx)
    source_ws = source_wb.active
    xls_sheet = xls_book.sheet_by_index(0)
    format_cells_as_text(source_ws)
    align_cells_left(source_ws)
        # 1. Gather the QTY value first
    QTY_total = 0
    row = 21
    column_length = 0


   # Calculate column length (assumes non-empty column B values starting from row 21)
    while True:
        try:
            value_from_B = xls_sheet.cell_value(row - 1, 1)  # B21 is (row 20, col 1) in zero-based indexing
            if value_from_B:
                try:
                    numeric_value = float(value_from_B)  # Convert the value to a float
                    QTY_total += numeric_value  # Add to QTY total
                    column_length += 1  # Count number of rows
                except (ValueError, TypeError):
                    print(f"Non-numeric value found in B{row}: {value_from_B}")  # Handle non-numeric values
            else:
                break  # End loop if no more values in column B
            row += 1
        except IndexError:
            break

    # Output the total QTY for debugging
    print(f"Total QTY: {QTY_total}")
    print(f"Column Length: {column_length}")

    # Mapping uploaded cells to copy cells
    data_map = {
        (15, 1): 'F3', (15, 4): 'F4', (15, 7): 'F5',
        (15, 9): 'F6', (15, 10): 'F7', (15, 11): 'F8'
    }

    for (row, col), copy_cell in data_map.items():
        value = xls_sheet.cell_value(row, col)
        source_ws[copy_cell] = value

    if QTY_total < 16:
        # Track numbers from A21 and below, copy to A20 in the copy
        row = 21
        copy_row = 14
        column_length = 0

        while True:
            try:
                value = xls_sheet.cell_value(row - 1, 0)  # Column A is index 0 (zero-based)
                if value:
                    source_ws[f'A{copy_row}'] = value
                    row += 1
                    copy_row += 1
                    column_length += 1  # Increment the length count
                else:
                    break
            except IndexError:
                break

        print(f"Length of Column A: {column_length}")  # Debugging print for column length

 
        manyToMany(xls_sheet, source_ws, 21, 7, 'F', 14, column_length)
        oneToMany(xls_sheet, source_ws, 3, 2, 'A', 14, column_length) # PO
        oneToMany(xls_sheet, source_ws, 15, 11, 'B', 14, column_length) #Zip code
        manyToMany(xls_sheet, source_ws, 21, 6, 'G', 14, column_length)
        manyToMany(xls_sheet, source_ws, 21, 8, 'H', 14, column_length) #SKU
        manyToMany(xls_sheet, source_ws, 21, 1, 'L', 14, column_length) #QTY
        manyToMany(xls_sheet, source_ws, 21, 1, 'M', 14, column_length) #Label #
        manyToMany(xls_sheet, source_ws, 21, 4, 'I', 14, column_length)


        # Now copy "mixed" into J14 and K14 down for the length of other items
        for i in range(14, 14 + column_length):  # Loop through rows starting at J14 and K14
            source_ws[f'J{i}'] = "NA"  # Paste "mixed" into J14 and down
            source_ws[f'K{i}'] = "NA"  # Paste "mixed" into K14 and down
            source_ws[f'E{i}'] = "NA"  # Paste "mixed" into K14 and down
            print(f"Pasting 'NA' into J{i}, E{i}, and K{i}")  # Debugging print

        # Now copy "mixed" into J14 and K14 down for the length of other items
        for i in range(14, 14 + column_length):  # Loop through rows starting at J14 and K14
            source_ws[f'C{i}'] = "Fedex"  # Paste "mixed" into J14 and down
            print(f"Pasting 'mixed' into C{i}")  # Debugging print

    else:
        data_map = {
            (3, 2): 'A14',  # C4 is (row 3, col 2) in zero-based indexing, copying to A14
            (15, 11): 'B14'  # L16 is (row 15, col 11) in zero-based indexing, copying to B14
        }

        for (row, col), copy_cell in data_map.items():
            value = xls_sheet.cell_value(row, col)
            source_ws[copy_cell] = value  # Paste value into the corresponding cell
            print(f"Pasting {value} from ({row + 1}, {col + 1}) to {copy_cell}")  # Debugging print
            # Set static values in specified cells
            source_ws['C14'] = "SAIA"
            source_ws['F14'] = "mixed"
            source_ws['G14'] = "mixed"
            source_ws['H14'] = "mixed"
            source_ws['I14'] = "mixed"
            source_ws['J14'] = "NA"
            source_ws['K14'] = "NA"
            source_ws['L14'] = "mixed"
            source_ws['M14'] = 1

            # Debugging print statements for the static values
            print("Pasting 'SAIA' into C14")
            print("Pasting 'mixed' into F14, G14, H14, I14, J14, K14")
            print("Pasting 1 into L14 and M14")


    format_cells_as_text(source_ws)
    align_cells_left(source_ws)

    # Save the updated copy
    source_wb.save(dest_file)

def process_ChewyLabel(file_path):
    """Main function to process Chewy UCC128 Label Request files."""
    current_date = datetime.datetime.now().strftime("%m.%d.%Y")

    # Determine if the file is XLSX or XLS and extract the PO number
    if file_path.endswith('.xlsx'):
        uploaded_wb = load_workbook(file_path)
        po_number = uploaded_wb.active['C4'].value
    elif file_path.endswith('.xls'):
        xls_book = xlrd.open_workbook(file_path)
        po_number = xls_book.sheet_by_index(0).cell_value(3, 2)

    # Define the backup file path in FINISHED_FOLDER with a Chewy subfolder
    backup_file = os.path.join(FINISHED_FOLDER, f"Chewy/Chewy UCC128 Label Request PO {po_number} {current_date}.xlsx")
    
    # Ensure the directory exists
    os.makedirs(os.path.dirname(backup_file), exist_ok=True)

    # Perform the copy or conversion based on file type
    if file_path.endswith('.xlsx'):
        copy_xlsx_data(file_path, backup_file)
    elif file_path.endswith('.xls'):
        convert_xls_data(file_path, backup_file)

    return backup_file, po_number
