"""
Pet Supermarket-specific calendar functionality and ship date calculations.
"""
import datetime
from openpyxl import load_workbook
import xlrd

def get_pet_supermarket_ship_date_advanced(file_path, **kwargs):
    """
    Advanced Pet Supermarket ship date calculation with additional business rules.
    
    Args:
        file_path: Path to the Pet Supermarket Excel file
        **kwargs: Additional parameters for future expansion
    
    Returns:
        datetime.datetime: Recommended ship date or None if no automatic date
    """
    try:
        if file_path.endswith('.xlsx'):
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Read due date from cell C7
            due_date_str = ws['C7'].value
        elif file_path.endswith('.xls'):
            xls_book = xlrd.open_workbook(file_path)
            xls_sheet = xls_book.sheet_by_index(0)
            
            # Read due date from cell C7 (row 6, col 2 in 0-based indexing)
            due_date_str = xls_sheet.cell_value(6, 2)
        
        # Parse the due date
        due_date = parse_date_string(due_date_str)
        
        if due_date:
            # Pet Supermarket business rules:
            # - Standard: 12 business days before due date
            # - Rush orders: 8 business days before due date (if special flag)
            # - Weekend adjustment: ensure ship date is not on weekend
            
            # Check for rush order indicators (could be expanded)
            is_rush = kwargs.get('rush_order', False)
            business_days = 8 if is_rush else 12
            
            ship_date = calculate_business_days(due_date, business_days)
            
            # Adjust for weekends if ship date falls on weekend
            while ship_date.weekday() >= 5:  # Saturday = 5, Sunday = 6
                ship_date -= datetime.timedelta(days=1)
            
            return ship_date
        else:
            return None
            
    except Exception as e:
        print(f"Error calculating Pet Supermarket ship date: {e}")
        return None

def get_pet_supermarket_order_details(file_path):
    """
    Extract additional order details from Pet Supermarket files.
    
    Args:
        file_path: Path to the Pet Supermarket Excel file
    
    Returns:
        dict: Order details including due date, order type, etc.
    """
    try:
        if file_path.endswith('.xlsx'):
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Read various cells for order details
            po_number = ws['C4'].value
            due_date_str = ws['C7'].value
            order_type = ws.get('D7', None)  # Could be used for order classification
            
        elif file_path.endswith('.xls'):
            xls_book = xlrd.open_workbook(file_path)
            xls_sheet = xls_book.sheet_by_index(0)
            
            po_number = xls_sheet.cell_value(3, 2)  # C4
            due_date_str = xls_sheet.cell_value(6, 2)  # C7
            try:
                order_type = xls_sheet.cell_value(6, 3)  # D7
            except IndexError:
                order_type = None
        
        due_date = parse_date_string(due_date_str)
        
        return {
            'po_number': po_number,
            'due_date': due_date,
            'order_type': order_type,
            'days_until_due': (due_date - datetime.datetime.now().date()).days if due_date else None
        }
        
    except Exception as e:
        print(f"Error extracting Pet Supermarket order details: {e}")
        return {}

def calculate_business_days(start_date, days):
    """Calculate a date that is a certain number of business days before the start date."""
    current_date = start_date
    business_days_counted = 0
    
    while business_days_counted < days:
        current_date -= datetime.timedelta(days=1)
        # Check if it's a weekday (Monday = 0, Sunday = 6)
        if current_date.weekday() < 5:  # Monday to Friday
            business_days_counted += 1
    
    return current_date

def parse_date_string(date_string):
    """Parse various date string formats and return a datetime object."""
    if not date_string:
        return None
    
    # Remove any extra whitespace
    date_string = str(date_string).strip()
    
    # Try different date formats
    date_formats = [
        '%m/%d/%Y',
        '%m/%d/%y',
        '%m-%d-%Y',
        '%m-%d-%y',
        '%Y-%m-%d',
        '%m/%d/%Y %H:%M:%S',
        '%m/%d/%y %H:%M:%S'
    ]
    
    for fmt in date_formats:
        try:
            return datetime.datetime.strptime(date_string, fmt)
        except ValueError:
            continue
    
    # If none of the formats work, try to extract date from Excel serial number
    try:
        # Excel serial number (days since 1900-01-01)
        excel_date = float(date_string)
        # Excel epoch is 1900-01-01, but Excel incorrectly treats 1900 as a leap year
        # So we need to adjust for dates after 1900-02-28
        if excel_date > 59:  # After 1900-02-28
            excel_date -= 1
        return datetime.datetime(1900, 1, 1) + datetime.timedelta(days=excel_date - 1)
    except (ValueError, TypeError):
        pass
    
    return None 