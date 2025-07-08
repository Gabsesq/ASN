"""
Chewy-specific calendar functionality and ship date calculations.
"""
import datetime
from openpyxl import load_workbook
import xlrd

def get_chewy_ship_date_advanced(file_path, **kwargs):
    """
    Advanced Chewy ship date calculation with additional business rules.
    
    Args:
        file_path: Path to the Chewy Excel file
        **kwargs: Additional parameters for future expansion
    
    Returns:
        datetime.datetime: Recommended ship date or None if no automatic date
    """
    try:
        if file_path.endswith('.xlsx'):
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Read quantities from column B starting at row 21
            quantities = []
            row = 21
            while ws[f'B{row}'].value is not None:
                try:
                    qty = int(ws[f'B{row}'].value)
                    quantities.append(qty)
                except (ValueError, TypeError):
                    pass
                row += 1
        elif file_path.endswith('.xls'):
            xls_book = xlrd.open_workbook(file_path)
            xls_sheet = xls_book.sheet_by_index(0)
            
            # Read quantities from column B starting at row 21
            quantities = []
            row = 20  # xlrd uses 0-based indexing
            while True:
                try:
                    qty = xls_sheet.cell_value(row, 1)  # Column B is index 1
                    if qty:
                        try:
                            qty = int(qty)
                            quantities.append(qty)
                        except (ValueError, TypeError):
                            pass
                    else:
                        break
                    row += 1
                except IndexError:
                    break
        
        # Calculate total quantity
        total_qty = sum(quantities)
        
        # Chewy business rules:
        # - If total quantity > 100: ship next day (FedEx)
        # - If total quantity <= 100: no automatic ship date (manual review needed)
        # - Special handling for weekends: if next day is weekend, ship Monday
        
        if total_qty > 100:
            ship_date = datetime.datetime.now() + datetime.timedelta(days=1)
            
            # Adjust for weekends
            while ship_date.weekday() >= 5:  # Saturday = 5, Sunday = 6
                ship_date += datetime.timedelta(days=1)
            
            return ship_date
        else:
            return None  # No automatic ship date for small orders
            
    except Exception as e:
        print(f"Error calculating Chewy ship date: {e}")
        return None

def get_chewy_order_priority(file_path):
    """
    Determine Chewy order priority based on file content.
    
    Args:
        file_path: Path to the Chewy Excel file
    
    Returns:
        str: Priority level ('high', 'medium', 'low')
    """
    try:
        if file_path.endswith('.xlsx'):
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Read quantities from column B starting at row 21
            quantities = []
            row = 21
            while ws[f'B{row}'].value is not None:
                try:
                    qty = int(ws[f'B{row}'].value)
                    quantities.append(qty)
                except (ValueError, TypeError):
                    pass
                row += 1
        elif file_path.endswith('.xls'):
            xls_book = xlrd.open_workbook(file_path)
            xls_sheet = xls_book.sheet_by_index(0)
            
            quantities = []
            row = 20  # xlrd uses 0-based indexing
            while True:
                try:
                    qty = xls_sheet.cell_value(row, 1)  # Column B is index 1
                    if qty:
                        try:
                            qty = int(qty)
                            quantities.append(qty)
                        except (ValueError, TypeError):
                            pass
                    else:
                        break
                    row += 1
                except IndexError:
                    break
        
        total_qty = sum(quantities)
        
        if total_qty > 200:
            return 'high'
        elif total_qty > 100:
            return 'medium'
        else:
            return 'low'
            
    except Exception as e:
        print(f"Error determining Chewy order priority: {e}")
        return 'low'

def get_chewy_location(file_path):
    """
    Get the Chewy location from cell B16.
    """
    try:
        if file_path.endswith('.xlsx'):
            wb = load_workbook(file_path)
            ws = wb.active
            location = ws['B16'].value
        elif file_path.endswith('.xls'):
            xls_book = xlrd.open_workbook(file_path)
            xls_sheet = xls_book.sheet_by_index(0)
            location = xls_sheet.cell_value(15, 1)
        return str(location).strip() if location else None
    except Exception as e:
        print(f"Error reading Chewy location: {e}")
        return None

def get_chewy_order_details(file_path):
    """
    Extract comprehensive order details from Chewy files.
    
    Args:
        file_path: Path to the Chewy Excel file
    
    Returns:
        dict: Order details including location, priority, quantities, etc.
    """
    try:
        if file_path.endswith('.xlsx'):
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Read quantities from column B starting at row 21
            quantities = []
            row = 21
            while ws[f'B{row}'].value is not None:
                try:
                    qty = int(ws[f'B{row}'].value)
                    quantities.append(qty)
                except (ValueError, TypeError):
                    pass
                row += 1
                
        elif file_path.endswith('.xls'):
            xls_book = xlrd.open_workbook(file_path)
            xls_sheet = xls_book.sheet_by_index(0)
            
            quantities = []
            row = 20  # xlrd uses 0-based indexing
            while True:
                try:
                    qty = xls_sheet.cell_value(row, 1)  # Column B is index 1
                    if qty:
                        try:
                            qty = int(qty)
                            quantities.append(qty)
                        except (ValueError, TypeError):
                            pass
                    else:
                        break
                    row += 1
                except IndexError:
                    break
        
        total_qty = sum(quantities)
        priority = get_chewy_order_priority(file_path)
        location = get_chewy_location(file_path)
        ship_date = get_chewy_ship_date_advanced(file_path)
        
        return {
            'location': location,
            'priority': priority,
            'total_quantity': total_qty,
            'ship_date': ship_date,
            'line_items': len(quantities),
            'requires_fedex': total_qty > 100
        }
        
    except Exception as e:
        print(f"Error extracting Chewy order details: {e}")
        return {}

def get_chewy_ship_date_next_day():
    """
    Always return the next day's date as the ship date.
    """
    import datetime
    ship_date = datetime.datetime.now() + datetime.timedelta(days=1)
    # Adjust for weekends: if next day is Sat/Sun, move to Monday
    while ship_date.weekday() >= 5:
        ship_date += datetime.timedelta(days=1)
    return ship_date

def get_chewy_event_details(file_path, po_number):
    """
    Return only location and next-day ship date for Chewy.
    """
    location = get_chewy_location(file_path)
    ship_date = get_chewy_ship_date_next_day()
    return {
        'location': location,
        'ship_date': ship_date,
        'po_number': po_number
    } 