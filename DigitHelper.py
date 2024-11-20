from openpyxl import load_workbook
import xlrd
import os


def process_single_label_with_description(
    asn_file_path,
    carton_label_file_path,
    config,
    output_file_name="Updated_ASN_SingleLabel.xlsx",
    mixed_description="Mixed"
):
    """
    Process ASN file with a single carton label for all rows.
    """
    # Load the ASN file
    if asn_file_path.endswith(".xls"):
        raise ValueError("The ASN file must be in .xlsx format. Please convert it.")
    asn_wb = load_workbook(asn_file_path)
    asn_ws = asn_wb.active

    # Extract the carton label from the carton label file
    if carton_label_file_path.endswith(".xls"):
        import xlrd
        xls_book = xlrd.open_workbook(carton_label_file_path)
        xls_sheet = xls_book.sheet_by_index(0)
        carton_label = xls_sheet.cell_value(config["start_row_label"] - 1, config["label_col"] - 1)
    elif carton_label_file_path.endswith(".xlsx"):
        carton_label_wb = load_workbook(carton_label_file_path)
        carton_label_ws = carton_label_wb.active
        carton_label = carton_label_ws.cell(row=config["start_row_label"], column=config["label_col"]).value
    else:
        raise ValueError("Unsupported file format for carton label file. Please upload a .xls or .xlsx file.")

    if not carton_label:
        raise ValueError("Carton label not found in the provided file.")

    print(f"Using single carton label: {carton_label}")

    # Determine the number of filled rows in Column A (or other column as configured)
    column_length = get_filled_rows_count(asn_ws, config["start_row_asn"], 1)
    print(f"Column length calculated: {column_length}")

    # Apply the single carton label and mixed description
    for row in range(config["start_row_asn"], config["start_row_asn"] + column_length):
        asn_ws.cell(row=row, column=config["label_col"]).value = carton_label

    # Save the updated file
    output_path = os.path.join(os.path.dirname(asn_file_path), output_file_name)
    asn_wb.save(output_path)
    print(f"Updated ASN file saved to: {output_path}")
    return output_path



def process_20_digit_labels(
    asn_file_path,
    label_file_path,
    config,
    output_file_name="Updated_ASN.xlsx"
):
    """
    Process ASN file with 20-digit carton labels for each line item.
    """
    # Load the ASN file
    if asn_file_path.endswith(".xls"):
        raise ValueError("The ASN file must be in .xlsx format. Please convert it.")
    asn_wb = load_workbook(asn_file_path)
    asn_ws = asn_wb.active

    # Parse the carton label file
    carton_data = []
    if label_file_path.endswith(".xls"):
        xls_book = xlrd.open_workbook(label_file_path)
        xls_sheet = xls_book.sheet_by_index(0)
        
        for row_idx in range(config["start_row_label"] - 1, xls_sheet.nrows):
            row = xls_sheet.row_values(row_idx)
            carton_label = str(row[config["label_col"] - 1]).strip() if row[config["label_col"] - 1] else None
            upc = str(row[config["upc_col"] - 1]).strip() if row[config["upc_col"] - 1] else None
            vendor_part = str(row[config["vendor_part_col"] - 1]).strip() if row[config["vendor_part_col"] - 1] else None
            sku = str(row[config["sku_col"] - 1]).strip() if row[config["sku_col"] - 1] else None

            if carton_label:
                carton_data.append({
                    "carton_label": carton_label,
                    "upc": upc,
                    "vendor_part": vendor_part,
                    "sku": sku,
                })
    elif label_file_path.endswith(".xlsx"):
        carton_label_wb = load_workbook(label_file_path)
        carton_label_ws = carton_label_wb.active

        for row in carton_label_ws.iter_rows(min_row=config["start_row_label"], values_only=True):
            carton_label = str(row[config["label_col"] - 1]).strip() if row[config["label_col"] - 1] else None
            upc = str(row[config["upc_col"] - 1]).strip() if row[config["upc_col"] - 1] else None
            vendor_part = str(row[config["vendor_part_col"] - 1]).strip() if row[config["vendor_part_col"] - 1] else None
            sku = str(row[config["sku_col"] - 1]).strip() if row[config["sku_col"] - 1] else None

            if carton_label:
                carton_data.append({
                    "carton_label": carton_label,
                    "upc": upc,
                    "vendor_part": vendor_part,
                    "sku": sku,
                })
    else:
        raise ValueError("Unsupported file format. Please upload a .xls or .xlsx file for the carton labels.")

    # Update the ASN file with the parsed carton labels
    used_labels = set()
    for asn_row in asn_ws.iter_rows(min_row=config["start_row_asn"], values_only=False):
        upc_cell, vendor_part_cell, sku_cell, label_cell = (
            asn_row[config["upc_col"] - 1],
            asn_row[config["vendor_part_col"] - 1],
            asn_row[config["sku_col"] - 1],
            asn_row[config["label_col"] - 1],
        )
        upc = str(upc_cell.value).strip() if upc_cell.value else None
        vendor_part = str(vendor_part_cell.value).strip() if vendor_part_cell.value else None
        sku = str(sku_cell.value).strip() if sku_cell.value else None

        if not (upc and vendor_part and sku):
            continue

        for carton in carton_data:
            if (
                carton["upc"] == upc and
                carton["vendor_part"] == vendor_part and
                carton["sku"] == sku and
                carton["carton_label"] not in used_labels
            ):
                label_cell.value = carton["carton_label"]
                used_labels.add(carton["carton_label"])
                break

    # Save the updated file
    output_path = os.path.join(os.path.dirname(asn_file_path), output_file_name)
    asn_wb.save(output_path)
    print(f"Updated ASN file saved to: {output_path}")
    return output_path



def get_filled_rows_count(sheet, start_row, start_col):
    """
    Calculate the number of non-empty rows starting from a specific row and column.

    Parameters:
        sheet: The Excel sheet object (openpyxl Worksheet).
        start_row (int): The row to start processing (1-based index).
        start_col (int): The column to check for non-empty cells (1-based index).

    Returns:
        int: The count of non-empty rows starting from the given row and column.
    """
    filled_rows = 0
    row = start_row

    print(f"Calculating filled rows starting from row {start_row}, column {start_col}...")

    while True:
        try:
            value = sheet.cell(row=row, column=start_col).value  # Get the cell value
            print(f"Row {row}: Value in Column {start_col} = '{value}'")

            if value:  # If the cell is not empty, increment the count
                filled_rows += 1
                row += 1
            else:  # Stop counting if an empty cell is found
                break
        except Exception as e:
            print(f"Error while accessing row {row}: {e}")
            break

    print(f"Total filled rows: {filled_rows}")
    return filled_rows
